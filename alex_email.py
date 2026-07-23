import os
import re
import json
import time
import imaplib
import email
import schedule
import logging
import io
from email.header import decode_header
from datetime import datetime, timedelta
from anthropic import Anthropic
import gspread
from google.oauth2.service_account import Credentials
import resend

ZOHO_EMAIL        = os.environ.get("ZOHO_EMAIL", "internal@scope-iq.io")
ZOHO_APP_PASSWORD = os.environ.get("ZOHO_APP_PASSWORD")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
RESEND_API_KEY    = os.environ.get("RESEND_API_KEY")
GOOGLE_SHEET_ID   = os.environ.get("GOOGLE_SHEET_ID", "1_4L63VqFN6etwLRWW2zT0WyJTHUT8LLTWwqiNPJJR4w")
GOOGLE_CREDS_FILE = "primordial-mile-495807-k9-0217981265dd.json"
MODEL             = "claude-haiku-4-5-20251001"

IMAP_HOST        = "imappro.zoho.com"
IMAP_PORT        = 993
IMAP_TIMEOUT     = 30
REMINDER_1_DAYS  = 3
REMINDER_2_DAYS  = 7
REMINDER_3_DAYS  = 14
AUTO_CLOSE_DAYS  = 21
ALWAYS_CC        = "alishir.aliyev@scopeconsulting.az"
INTERNAL_DOMAIN  = os.environ.get("INTERNAL_DOMAIN", "scopeconsulting.az")
CLIENT_DOMAINS   = [d.strip().lower() for d in os.environ.get("CLIENT_DOMAINS", "").split(",") if d.strip()]
SCOPE_TEAM_EMAILS = [e.strip().lower() for e in os.environ.get("SCOPE_TEAM_EMAILS", "").split(",") if e.strip()]
REPORT_RECIPIENTS = [
    "alishir.aliyev@scopeconsulting.az",
    "afgan.mammadov@scopeconsulting.az",
    "techoffice@scopeconsulting.az"
]

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

anthropic_client     = Anthropic(api_key=ANTHROPIC_API_KEY)
_processed_ids_cache = set()
_cache_loaded        = False
_file_memory_cache   = {}

SYSTEM_PROMPT = """You are Alex Rivera, Construction Expert at SCOPE Consulting MMC.

You have 25 years of experience across commercial, residential, infrastructure, oil and gas, and high-end fit-out projects in Europe, the Middle East, and CIS countries including Azerbaijan.

LANGUAGE RULES:
Default language is English. Reply in English unless the incoming email is entirely in Azerbaijani (use correct special characters: ə ı ö ü ğ ş ç). Never mix languages.

EMAIL WRITING STYLE - CRITICAL:
Plain professional prose only. No bullet points, symbols, emojis, decorative dashes.
Formal salutation and closing. Direct, confident, senior tone.

RESPONSE TIME - CRITICAL:
Never say you will review later. Always analyse immediately.

SIGNATURE — always end every email with exactly this:

Alex Rivera
Construction Expert
SCOPE Consulting MMC
internal@scope-iq.io"""


def extract_emails_from_text(text):
    if not text:
        return []
    pattern = r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}'
    return list(set(m.lower() for m in re.findall(pattern, text)))


def get_imap_connection():
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    mail.socket().settimeout(IMAP_TIMEOUT)
    mail.login(ZOHO_EMAIL, ZOHO_APP_PASSWORD)
    return mail


def safe_logout(mail):
    try:
        mail.logout()
    except Exception:
        pass


def strip_quoted_reply(body):
    if not body:
        return body
    lines = body.split("\n")
    cut_index = len(lines)
    for i, line in enumerate(lines):
        stripped = line.strip().lower()
        if not stripped:
            continue
        if stripped.startswith(">"):
            cut_index = i; break
        if stripped.startswith("on ") and "wrote:" in stripped:
            cut_index = i; break
        if stripped.startswith("-----original message-----"):
            cut_index = i; break
        if stripped in ["from:", "sent:", "subject:", "to:"]:
            cut_index = i; break
    result = "\n".join(lines[:cut_index]).strip()
    return result if result else body


def is_short_reply(body_clean):
    if not body_clean:
        return True
    return len(body_clean.split()) <= 15


def is_daily_report_reply(subject):
    return "daily report" in subject.lower()


def get_gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if creds_json:
        creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=scopes)
    else:
        creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
    return gspread.authorize(creds)


def get_sheet(tab="Sheet1"):
    try:
        client = get_gspread_client()
        return client.open_by_key(GOOGLE_SHEET_ID).worksheet(tab)
    except Exception as e:
        logger.error(f"Sheet error: {e}")
        return None


def get_or_create_sheet(title, rows=1000, cols=5):
    try:
        client = get_gspread_client()
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        try:
            return spreadsheet.worksheet(title)
        except:
            return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)
    except Exception as e:
        logger.error(f"Sheet create error ({title}): {e}")
        return None


def get_action_tracker_sheet():
    try:
        client = get_gspread_client()
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        try:
            sheet = spreadsheet.worksheet("Action Tracker")
            headers = sheet.row_values(1)
            for col, name in [(15, "All Thread Participants"), (16, "Client Emails"), (17, "Responsible Name")]:
                if len(headers) < col:
                    sheet.update_cell(1, col, name)
            return sheet
        except:
            sheet = spreadsheet.add_worksheet(title="Action Tracker", rows=2000, cols=17)
            sheet.append_row([
                "Date Logged", "Meeting Reference", "Action Item", "Responsible Party", "Responsible Email",
                "Due Date", "Status", "Last Reminded", "Reminder Count", "Thread ID", "MOM Sender",
                "Draft Sent", "External Reply", "Notes", "All Thread Participants", "Client Emails", "Responsible Name"
            ])
            return sheet
    except Exception as e:
        logger.error(f"Action tracker error: {e}")
        return None


def get_ncr_tracker_sheet():
    try:
        client = get_gspread_client()
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        try:
            sheet = spreadsheet.worksheet("NCR Tracker")
            headers = sheet.row_values(1)
            if len(headers) < 17:
                sheet.update_cell(1, 17, "Client Emails")
            return sheet
        except:
            sheet = spreadsheet.add_worksheet(title="NCR Tracker", rows=2000, cols=17)
            sheet.append_row([
                "Date Logged", "NCR Number", "Description", "Contractor", "Contractor Email", "Date Raised",
                "Status", "Last Reminded", "Reminder Count", "Thread ID", "Raised By", "CAR Received",
                "CAR Content", "Notes", "All Thread Participants", "Responsible Name", "Client Emails"
            ])
            return sheet
    except Exception as e:
        logger.error(f"NCR tracker error: {e}")
        return None


def save_files_to_memory(sender, attachments):
    global _file_memory_cache
    sender = sender.strip().lower()
    _file_memory_cache[sender] = [{"name": a["name"], "content": a["content"], "saved_at": datetime.now().strftime("%d.%m.%Y %H:%M")} for a in attachments]
    try:
        sheet = get_or_create_sheet("File Memory", rows=2000, cols=4)
        if not sheet:
            return
        all_values = sheet.get_all_values()
        if not all_values:
            sheet.append_row(["Sender", "Filename", "Content", "Saved At"])
            all_values = [["Sender", "Filename", "Content", "Saved At"]]
        rows_to_delete = [i + 1 for i, row in enumerate(all_values) if i > 0 and row and row[0].strip().lower() == sender]
        for r in reversed(rows_to_delete):
            sheet.delete_rows(r)
        saved_at = datetime.now().strftime("%d.%m.%Y %H:%M")
        for att in attachments:
            sheet.append_row([sender, att["name"], att["content"][:10000], saved_at])
    except Exception as e:
        logger.error(f"Save files error: {e}")


def load_files_from_memory(sender):
    global _file_memory_cache
    sender = sender.strip().lower()
    if sender in _file_memory_cache and _file_memory_cache[sender]:
        return _file_memory_cache[sender]
    try:
        sheet = get_or_create_sheet("File Memory", rows=2000, cols=4)
        if not sheet:
            return []
        files = []
        for row in sheet.get_all_values()[1:]:
            if row and len(row) >= 3 and row[0].strip().lower() == sender:
                files.append({"name": row[1], "content": row[2], "saved_at": row[3] if len(row) > 3 else ""})
        if files:
            _file_memory_cache[sender] = files
        return files
    except Exception as e:
        logger.error(f"Load files error: {e}")
        return []


def get_processed_sheet():
    return get_or_create_sheet("Processed Emails", rows=5000, cols=2)


def load_processed_ids():
    global _processed_ids_cache, _cache_loaded
    if _cache_loaded:
        return _processed_ids_cache
    try:
        sheet = get_processed_sheet()
        if sheet:
            all_values = sheet.get_all_values()
            if not all_values:
                sheet.append_row(["Message ID", "Processed At"])
            else:
                for row in all_values[1:]:
                    if row and row[0] and row[0].strip():
                        _processed_ids_cache.add(row[0].strip())
    except Exception as e:
        logger.error(f"Load IDs error: {e}")
    _cache_loaded = True
    return _processed_ids_cache


def mark_as_processed(msg_id):
    global _processed_ids_cache
    if not msg_id:
        return
    msg_id = str(msg_id).strip()
    if msg_id in _processed_ids_cache:
        return
    _processed_ids_cache.add(msg_id)
    try:
        sheet = get_processed_sheet()
        if sheet:
            sheet.append_row([msg_id, datetime.now().strftime("%d.%m.%Y %H:%M")])
    except Exception as e:
        logger.error(f"Mark processed error: {e}")


def is_processed(msg_id):
    if not msg_id:
        return False
    return str(msg_id).strip() in _processed_ids_cache


def save_to_monitoring(sender, subject, summary, action, thread_id, status="Monitoring"):
    try:
        sheet = get_sheet("Sheet1")
        if sheet:
            headers = sheet.row_values(1)
            for col, name in [(7, "Thread-ID"), (8, "Last Reminded"), (9, "Reminder Count")]:
                if len(headers) < col:
                    sheet.update_cell(1, col, name)
            sheet.append_row([datetime.now().strftime("%d.%m.%Y %H:%M"), sender, subject, summary, action, status, thread_id, "", "0"])
    except Exception as e:
        logger.error(f"Save monitoring error: {e}")


def save_to_memory(sender, subject, summary, action, status="Open"):
    try:
        sheet = get_sheet("Sheet1")
        if sheet:
            sheet.append_row([datetime.now().strftime("%d.%m.%Y %H:%M"), sender, subject, summary, action, status, "", "", "0"])
    except Exception as e:
        logger.error(f"Save error: {e}")


def update_row(row_number, status=None, last_reminded=None, reminder_count=None):
    try:
        sheet = get_sheet("Sheet1")
        if sheet:
            if status: sheet.update_cell(row_number, 6, status)
            if last_reminded: sheet.update_cell(row_number, 8, last_reminded)
            if reminder_count is not None: sheet.update_cell(row_number, 9, str(reminder_count))
    except Exception as e:
        logger.error(f"Update row error: {e}")


def save_action_item(meeting_ref, action_item, responsible_party, responsible_email, responsible_name, due_date,
                     thread_id, mom_sender, all_participants, client_emails, status="Open"):
    try:
        sheet = get_action_tracker_sheet()
        if sheet:
            sheet.append_row([
                datetime.now().strftime("%d.%m.%Y %H:%M"), meeting_ref, action_item, responsible_party,
                responsible_email, due_date, status, "", "0", thread_id, mom_sender, "", "", "",
                ",".join(all_participants), ",".join(client_emails), responsible_name
            ])
    except Exception as e:
        logger.error(f"Save action error: {e}")


def update_action_row(row_number, status=None, last_reminded=None, reminder_count=None, draft_sent=None, external_reply=None, notes=None):
    try:
        sheet = get_action_tracker_sheet()
        if sheet:
            if status: sheet.update_cell(row_number, 7, status)
            if last_reminded: sheet.update_cell(row_number, 8, last_reminded)
            if reminder_count is not None: sheet.update_cell(row_number, 9, str(reminder_count))
            if draft_sent: sheet.update_cell(row_number, 12, draft_sent)
            if external_reply: sheet.update_cell(row_number, 13, external_reply[:500])
            if notes: sheet.update_cell(row_number, 14, notes)
    except Exception as e:
        logger.error(f"Update action row error: {e}")


def save_ncr_item(ncr_number, description, contractor, contact_emails, contact_names, date_raised, thread_id, raised_by, all_participants, client_emails, status="Open"):
    try:
        sheet = get_ncr_tracker_sheet()
        if sheet:
            sheet.append_row([
                datetime.now().strftime("%d.%m.%Y %H:%M"), ncr_number, description, contractor,
                ",".join(contact_emails), date_raised, status, "", "0", thread_id, raised_by, "", "", "",
                ",".join(all_participants), ",".join(contact_names), ",".join(client_emails)
            ])
    except Exception as e:
        logger.error(f"Save NCR error: {e}")


def update_ncr_row(row_number, status=None, last_reminded=None, reminder_count=None, car_received=None, car_content=None, notes=None):
    try:
        sheet = get_ncr_tracker_sheet()
        if sheet:
            if status: sheet.update_cell(row_number, 7, status)
            if last_reminded: sheet.update_cell(row_number, 8, last_reminded)
            if reminder_count is not None: sheet.update_cell(row_number, 9, str(reminder_count))
            if car_received: sheet.update_cell(row_number, 12, car_received)
            if car_content: sheet.update_cell(row_number, 13, car_content[:500])
            if notes: sheet.update_cell(row_number, 14, notes)
    except Exception as e:
        logger.error(f"Update NCR row error: {e}")


def get_ncr_data_from_row(row):
    participants_raw = row[14].strip() if len(row) > 14 else ""
    all_participants = [p.strip() for p in participants_raw.split(",") if p.strip() and "@" in p.strip()]
    emails_raw = row[4].strip() if len(row) > 4 else ""
    all_emails = [e.strip() for e in emails_raw.split(",") if e.strip() and e.strip() != "UNKNOWN"]
    names_raw = row[15].strip() if len(row) > 15 else ""
    all_names = [n.strip() for n in names_raw.split(",") if n.strip()]
    client_raw = row[16].strip() if len(row) > 16 else ""
    client_emails = [c.strip() for c in client_raw.split(",") if c.strip() and "@" in c.strip()]
    return {
        "date": row[0].strip() if row[0] else "", "ncr_number": row[1].strip() if len(row) > 1 else "",
        "description": row[2].strip() if len(row) > 2 else "", "contractor": row[3].strip() if len(row) > 3 else "",
        "email": all_emails[0] if all_emails else "UNKNOWN", "all_emails": all_emails,
        "date_raised": row[5].strip() if len(row) > 5 else "", "status": row[6].strip() if len(row) > 6 else "",
        "last_reminded": row[7].strip() if len(row) > 7 else "",
        "reminder_count": int(row[8].strip()) if len(row) > 8 and row[8].strip().isdigit() else 0,
        "thread_id": row[9].strip() if len(row) > 9 else "", "raised_by": row[10].strip() if len(row) > 10 else "",
        "all_participants": all_participants, "responsible_name": all_names[0] if all_names else "",
        "all_names": all_names, "client_emails": client_emails
    }


def read_memory_for_report():
    try:
        sheet = get_sheet("Sheet1")
        if sheet:
            records = sheet.get_all_records()
            pending = [r for r in records if r.get("Status") in ["Open", "Monitoring"]]
            closed = [r for r in records if r.get("Status") in ["Closed", "Closed — No Response"]]
            return pending, closed
        return [], []
    except Exception as e:
        logger.error(f"Read memory error: {e}")
        return [], []


def read_actions_for_report():
    try:
        sheet = get_action_tracker_sheet()
        if sheet:
            records = sheet.get_all_records()
            open_actions = [r for r in records if r.get("Status") in ["Open", "Draft Pending", "Draft Sent", "Reminded"]]
            closed = [r for r in records if r.get("Status") in ["Closed", "Closed — No Response", "Closed — No Action Required"]]
            flagged = [r for r in records if r.get("Status") == "Email Unknown"]
            return open_actions, closed, flagged
        return [], [], []
    except Exception as e:
        logger.error(f"Read actions error: {e}")
        return [], [], []


def read_ncrs_for_report():
    try:
        sheet = get_ncr_tracker_sheet()
        if sheet:
            records = sheet.get_all_records()
            open_ncrs = [r for r in records if r.get("Status") in ["Open", "Reminded", "Email Unknown"]]
            closed_ncrs = [r for r in records if r.get("Status") == "Closed"]
            return open_ncrs, closed_ncrs
        return [], []
    except Exception as e:
        logger.error(f"Read NCRs error: {e}")
        return [], []


def get_first_name(email_address):
    try:
        local = email_address.split("@")[0]
        parts = local.replace(".", " ").replace("_", " ").split()
        return parts[0].capitalize() if parts else "Colleague"
    except:
        return "Colleague"


def is_internal_email(email_addr):
    return INTERNAL_DOMAIN.lower() in email_addr.lower()


def is_client_email(email_addr):
    if not CLIENT_DOMAINS:
        return False
    return any(d in email_addr.lower() for d in CLIENT_DOMAINS)


def is_approval_reply(body_text):
    kws = ["approve", "approved", "send", "go ahead", "ok", "okay", "confirmed", "confirm", "yes", "proceed", "looks good", "please send", "send it", "agreed"]
    return any(kw in body_text.lower() for kw in kws)


def is_rejection_reply(body_text):
    kws = ["no need", "not needed", "not required", "no longer required", "close this", "close it", "please close",
           "disregard", "ignore this", "not necessary", "cancel this", "cancel it", "reject", "no action required",
           "not applicable", "n/a", "drop this", "stop this", "no longer relevant", "obsolete", "already resolved",
           "already closed", "already handled", "resolved on site", "resolved verbally", "handled on site",
           "handled verbally", "closed on site", "sorted on site", "done on site", "agreed on site", "settled on site",
           "settled verbally", "resolved offline", "handled offline", "resolved directly", "no follow-up needed",
           "no followup needed", "no further action", "stop chasing", "stop following up", "verbally agreed",
           "verbally resolved", "verbally confirmed on site"]
    text = body_text.lower()
    return any(kw in text for kw in kws)


def is_mom_email(subject, body, attachments):
    kws = ["minutes of meeting", "mom", "meeting minutes", "iclasın protokolu", "görüş protokolu", "meeting notes", "action items", "action points", "minutes from", "meeting summary"]
    text = (subject + " " + body).lower()
    if any(kw in text for kw in kws):
        return True
    for att in attachments:
        if any(kw in att.get("name", "").lower() for kw in ["mom", "minutes", "meeting", "protocol"]):
            return True
    return False


def is_ncr_email(subject, body, attachments):
    kws = ["ncr", "non-conformance report", "non conformance report", "non-conformance", "non conformance", "corrective action request", "nonconformance"]
    text = (subject + " " + body).lower()
    if any(kw in text for kw in kws):
        return True
    for att in attachments:
        if any(kw in att.get("name", "").lower() for kw in ["ncr", "nonconformance", "non-conformance"]):
            return True
    return False


def extract_display_name(header_value):
    if not header_value:
        return ""
    header_str = safe_decode(header_value)
    if "<" in header_str:
        name = header_str[:header_str.rfind("<")].strip()
        return name.strip('"').strip("'").strip()
    return ""


def extract_email_with_name(header_value):
    return {"email": extract_email_address(header_value), "name": extract_display_name(header_value)}


def extract_all_emails_with_names(header_value):
    if not header_value:
        return []
    header_str = safe_decode(header_value)
    results, parts, current, in_quote = [], [], "", False
    for char in header_str:
        if char == '"':
            in_quote = not in_quote
        if char == "," and not in_quote:
            parts.append(current.strip()); current = ""
        else:
            current += char
    if current.strip():
        parts.append(current.strip())
    for part in parts:
        info = extract_email_with_name(part)
        if info["email"] and "@" in info["email"]:
            results.append(info)
    return results


def classify_thread_participants(all_participants_with_names):
    internal, clients, contractors = [], [], []
    for p in all_participants_with_names:
        addr = p["email"]
        if addr == ZOHO_EMAIL.lower():
            continue
        if is_internal_email(addr):
            internal.append(p)
        elif is_client_email(addr):
            clients.append(p)
        else:
            contractors.append(p)
    return internal, clients, contractors


def merge_party_classification(all_thread_with_names, claude_classification):
    email_to_entry = {c.get("email", "").lower(): c for c in claude_classification if c.get("email")}
    internal, dom_clients, dom_others = classify_thread_participants(all_thread_with_names)
    dom_client_emails = {p["email"].lower() for p in dom_clients}
    clients, others = [], []
    for p in all_thread_with_names:
        addr = p["email"]
        if addr == ZOHO_EMAIL.lower() or is_internal_email(addr):
            continue
        entry = email_to_entry.get(addr.lower())
        if entry:
            role_label = entry.get("role_label", "Other") or "Other"
            if entry.get("is_client"):
                clients.append({**p, "role_label": "Client"})
            else:
                others.append({**p, "role_label": role_label})
        else:
            if addr.lower() in dom_client_emails:
                clients.append({**p, "role_label": "Client"})
            else:
                others.append({**p, "role_label": "Contractor"})
    return internal, clients, others


def find_action_by_thread(in_reply_to, references):
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return None
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            thread_id = row[9].strip() if len(row) > 9 else ""
            status = row[6].strip() if len(row) > 6 else ""
            if not thread_id or status in ["Closed", "Closed — No Response", "Closed — No Action Required"]:
                continue
            if thread_id in in_reply_to or thread_id in references:
                return {"row": i, **get_action_data_from_row(row)}
        return None
    except Exception as e:
        logger.error(f"Find action error: {e}")
        return None


def find_all_actions_by_thread(thread_id):
    matches = []
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return matches
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            tid = row[9].strip() if len(row) > 9 else ""
            status = row[6].strip() if len(row) > 6 else ""
            if tid == thread_id and status not in ["Closed", "Closed — No Response", "Closed — No Action Required"]:
                matches.append({"row": i, **get_action_data_from_row(row)})
    except Exception as e:
        logger.error(f"Find all actions error: {e}")
    return matches


def find_all_open_actions_matching_refs(in_reply_to, references):
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return []
        results = []
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            thread_id = row[9].strip() if len(row) > 9 else ""
            status = row[6].strip() if len(row) > 6 else ""
            if not thread_id:
                continue
            if thread_id in in_reply_to or thread_id in references:
                if status not in ["Closed", "Closed — No Response", "Closed — No Action Required"]:
                    results.append({"row": i, **get_action_data_from_row(row)})
        return results
    except Exception as e:
        logger.error(f"Find matching refs error: {e}")
        return []


def find_open_actions_by_subject(subject):
    if not subject:
        return []
    clean_subject = subject.lower()
    for prefix in ["re:", "fwd:", "fw:"]:
        clean_subject = clean_subject.replace(prefix, "")
    clean_subject = clean_subject.strip()
    if not clean_subject:
        return []
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return []
        matches = []
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            meeting_ref = row[1].strip() if len(row) > 1 else ""
            status = row[6].strip() if len(row) > 6 else ""
            thread_id = row[9].strip() if len(row) > 9 else ""
            if not meeting_ref or not thread_id or status in ["Closed", "Closed — No Response", "Closed — No Action Required"]:
                continue
            if meeting_ref.lower() in clean_subject or clean_subject in meeting_ref.lower():
                matches.append({"row": i, **get_action_data_from_row(row)})
        return matches
    except Exception as e:
        logger.error(f"Subject fallback match error: {e}")
        return []


def find_all_open_ncrs_matching_refs(in_reply_to, references):
    try:
        sheet = get_ncr_tracker_sheet()
        if not sheet:
            return []
        results = []
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            thread_id = row[9].strip() if len(row) > 9 else ""
            status = row[6].strip() if len(row) > 6 else ""
            if not thread_id:
                continue
            if thread_id in in_reply_to or thread_id in references:
                if status != "Closed":
                    results.append({"row": i, **get_ncr_data_from_row(row)})
        return results
    except Exception as e:
        logger.error(f"Find NCR matching refs error: {e}")
        return []


def find_open_ncrs_by_subject(subject):
    if not subject:
        return []
    clean_subject = subject.lower()
    for prefix in ["re:", "fwd:", "fw:"]:
        clean_subject = clean_subject.replace(prefix, "")
    clean_subject = clean_subject.strip()
    if not clean_subject:
        return []
    try:
        sheet = get_ncr_tracker_sheet()
        if not sheet:
            return []
        matches = []
        for i, row in enumerate(sheet.get_all_values()[1:], start=2):
            if len(row) < 10:
                continue
            ncr_number = row[1].strip() if len(row) > 1 else ""
            status = row[6].strip() if len(row) > 6 else ""
            thread_id = row[9].strip() if len(row) > 9 else ""
            if not ncr_number or not thread_id or status == "Closed":
                continue
            if ncr_number.lower() in clean_subject or clean_subject in ncr_number.lower():
                matches.append({"row": i, **get_ncr_data_from_row(row)})
        return matches
    except Exception as e:
        logger.error(f"NCR subject fallback error: {e}")
        return []


def find_all_recent_open_ncrs():
    try:
        sheet = get_ncr_tracker_sheet()
        if not sheet:
            return []
        all_values = sheet.get_all_values()
        today = datetime.now()
        matches = []
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) < 10:
                continue
            status = row[6].strip() if len(row) > 6 else ""
            date_str = row[0].strip() if row[0] else ""
            if status != "Open":
                continue
            try:
                logged_date = datetime.strptime(date_str, "%d.%m.%Y %H:%M")
                if (today - logged_date).days > 3:
                    continue
            except:
                continue
            matches.append({"row": i, **get_ncr_data_from_row(row)})
        return matches
    except Exception as e:
        logger.error(f"Find recent open NCRs error: {e}")
        return []


def extract_mom_actions(mom_content, thread_participants_with_names, subject):
    try:
        external_participants = [p for p in thread_participants_with_names if p["email"] != ZOHO_EMAIL.lower() and not is_internal_email(p["email"])]

        def fmt_list(lst):
            return "\n".join([f"  - {p['name']} <{p['email']}>" if p['name'] else f"  - {p['email']}" for p in lst]) or "  None detected"

        context = f"External parties present in this email thread (To/CC), not yet classified:\n{fmt_list(external_participants)}"
        prompt = f"""Analyse this Minutes of Meeting and extract all action items.

Meeting subject: {subject}

{context}

MOM Content:
{mom_content[:16000]}

IMPORTANT — PARTY IDENTIFICATION:
The MOM document itself may explicitly label parties, for example "Client: Pasha Bank", "Employer: ...", "Contractor: ...", "Consultant: ...", "Designer: ...", "Building Operator: ...", "Subcontractor: ...", "Supplier: ...", "PMC: ...". Search the document text carefully for every such label. Match each labelled company name to the external email addresses listed above by comparing the company name to the email domain. For every external email listed above, identify its role_label using the exact role wording from the document. If not labelled, infer from context. Mark is_client true only for Client/Employer/Owner.

For each action item:
1. Action description
2. Responsible party name
3. Responsible email — match to external parties above. If no confident match write UNKNOWN.
4. Responsible display name
5. Due date or NOT SPECIFIED
6. Meeting reference
7. Responsible role label (Contractor, Designer, Building Operator, Consultant, etc.) or SCOPE

Respond in this exact JSON only:
{{"meeting_reference": "...", "client_identified": "...", "contractor_identified": "...", "party_classification": [{{"email": "...", "role_label": "...", "is_client": true or false, "company": "..."}}], "actions": [{{"action": "...", "responsible_party": "...", "responsible_email": "...", "responsible_name": "...", "due_date": "...", "responsible_role_label": "..."}}]}}"""

        response = anthropic_client.messages.create(model=MODEL, max_tokens=2500, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        return json.loads(text.strip())
    except Exception as e:
        logger.error(f"MOM extraction error: {e}")
        return {"meeting_reference": subject, "actions": [], "client_identified": "Unknown", "contractor_identified": "Unknown", "party_classification": []}


def extract_ncr_details(ncr_content, thread_participants_with_names, subject):
    try:
        external_participants = [p for p in thread_participants_with_names if p["email"] != ZOHO_EMAIL.lower() and not is_internal_email(p["email"])]

        def fmt_list(lst):
            return "\n".join([f"  - {p['name']} <{p['email']}>" if p['name'] else f"  - {p['email']}" for p in lst]) or "  None detected"

        prompt = f"""Analyse this Non-Conformance Report (NCR).

Subject: {subject}

External parties in this email thread (ONLY valid candidates for contractor contact — never an internal SCOPE Consulting address):
{fmt_list(external_participants)}

NCR Content:
{ncr_content[:12000]}

IMPORTANT: The contractor is always an EXTERNAL party — never SCOPE Consulting MMC. SCOPE is the PMC raising the NCR.

Extract:
1. NCR number or reference exactly as written
2. Description of the non-conformance
3. Responsible contractor company name
4. All relevant contractor contacts — every external person/email belonging to the responsible contractor (there may be several). For each provide name and email.
5. Date raised as written, or NOT SPECIFIED

Respond in this exact JSON only:
{{"ncr_number": "...", "description": "...", "contractor": "...", "contacts": [{{"name": "...", "email": "..."}}], "date_raised": "..."}}"""

        response = anthropic_client.messages.create(model=MODEL, max_tokens=1200, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        data = json.loads(text.strip())
        data["contacts"] = [c for c in data.get("contacts", []) if c.get("email") and not is_internal_email(c["email"])]
        return data
    except Exception as e:
        logger.error(f"NCR extraction error: {e}")
        return {"ncr_number": "Unknown", "description": subject, "contractor": "Unknown", "contacts": [], "date_raised": "Not specified"}


def analyse_external_reply(action_item, reply_content):
    try:
        prompt = f"""Review this external party reply against a required action item.

Action required: {action_item}
Reply received: {reply_content[:3000]}

Has this reply fully and satisfactorily addressed the action?

Respond in this exact JSON only:
{{"satisfied": true or false, "analysis": "2-3 sentence assessment.", "outstanding_items": "What remains or NONE"}}"""
        response = anthropic_client.messages.create(model=MODEL, max_tokens=500, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        return json.loads(text.strip())
    except Exception as e:
        logger.error(f"Reply analysis error: {e}")
        return {"satisfied": False, "analysis": "Unable to analyse.", "outstanding_items": "Unknown"}


def extract_mom_clarification(reply_body, actions_summary):
    try:
        prompt = f"""An internal team member replied to a MOM action item confirmation with clarifications.

Current action items:
{actions_summary}

Reply received:
{reply_body[:3000]}

Extract any party clarifications — names/companies paired with an email, or statements identifying client/contractor/designer/etc. Provide party_name, email, role.

Respond in this exact JSON only:
{{"clarifications": [{{"party_name": "...", "email": "...", "role": "..."}}], "acknowledgement_note": "one sentence or NONE"}}"""
        response = anthropic_client.messages.create(model=MODEL, max_tokens=500, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        return json.loads(text.strip())
    except Exception as e:
        logger.error(f"Clarification parse error: {e}")
        return {"clarifications": [], "acknowledgement_note": "NONE"}


def extract_ncr_clarification(reply_body, ncr_summary):
    try:
        prompt = f"""An internal team member replied to an NCR notification with clarifications about contractor contacts.

Current NCR(s):
{ncr_summary}

Reply received:
{reply_body[:3000]}

Extract every contractor contact name and email mentioned — only external contacts, never internal SCOPE addresses.

Respond in this exact JSON only:
{{"contacts": [{{"name": "...", "email": "..."}}]}}"""
        response = anthropic_client.messages.create(model=MODEL, max_tokens=500, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        data = json.loads(text.strip())
        return [c for c in data.get("contacts", []) if c.get("email") and not is_internal_email(c["email"])]
    except Exception as e:
        logger.error(f"NCR clarification parse error: {e}")
        return []


def extract_daily_report_requests(body, open_actions, open_ncrs):
    try:
        def fmt_actions(lst):
            return "\n".join([f"  - ACTION_REF:{a['row']} | {a['meeting_ref']} | {a['action'][:100]}" for a in lst]) or "  None"

        def fmt_ncrs(lst):
            return "\n".join([f"  - NCR_REF:{n['row']} | {n['ncr_number']} | {n['description'][:100]}" for n in lst]) or "  None"

        prompt = f"""A team member replied to the SCOPE IQ Daily Report email with an instruction.

Open MOM actions:
{fmt_actions(open_actions)}

Open NCRs:
{fmt_ncrs(open_ncrs)}

Reply received:
{body[:2000]}

Identify which item(s) the person is referring to, using the ACTION_REF or NCR_REF identifiers above, and what they want done — "reminder" (send a follow-up/chase) or "close" (no longer needed). A reply can reference multiple items.

Respond in this exact JSON only:
{{"requests": [{{"ref_type": "ACTION" or "NCR", "ref": <row number as integer>, "request": "reminder" or "close"}}]}}"""

        response = anthropic_client.messages.create(model=MODEL, max_tokens=600, messages=[{"role": "user", "content": prompt}])
        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
        data = json.loads(text.strip())
        return data.get("requests", [])
    except Exception as e:
        logger.error(f"Daily report request parse error: {e}")
        return []


def append_to_action_list_column(row_number, col, value):
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return
        current = sheet.cell(row_number, col).value or ""
        items = [x.strip() for x in current.split(",") if x.strip()]
        if value.lower() not in [x.lower() for x in items]:
            items.append(value)
        sheet.update_cell(row_number, col, ",".join(items))
    except Exception as e:
        logger.error(f"Append list column error: {e}")


def update_action_contact(row_number, responsible_email=None, responsible_name=None, status=None):
    try:
        sheet = get_action_tracker_sheet()
        if sheet:
            if responsible_email: sheet.update_cell(row_number, 5, responsible_email)
            if responsible_name: sheet.update_cell(row_number, 17, responsible_name)
            if status: sheet.update_cell(row_number, 7, status)
    except Exception as e:
        logger.error(f"Update action contact error: {e}")


def apply_mom_clarifications(thread_actions, clarifications):
    updated = []
    for c in clarifications:
        name = (c.get("party_name") or "").strip()
        email_ = (c.get("email") or "").strip()
        role = (c.get("role") or "").strip().upper()
        if not email_ or "@" not in email_:
            continue
        if role == "CLIENT":
            for a in thread_actions:
                append_to_action_list_column(a["row"], 16, email_)
                append_to_action_list_column(a["row"], 15, email_)
            updated.append(f"Client contact recorded: {name} <{email_}>" if name else f"Client contact recorded: {email_}")
            continue
        matched_any = False
        for a in thread_actions:
            party = (a.get("responsible") or "").lower()
            if a["status"] == "Email Unknown" and name and (name.lower() in party or party in name.lower()):
                update_action_contact(a["row"], responsible_email=email_, responsible_name=name, status="Open")
                append_to_action_list_column(a["row"], 15, email_)
                updated.append(f"Updated '{a['action'][:60]}' — responsible: {name} <{email_}>")
                matched_any = True
        if not matched_any:
            unknown_rows = [a for a in thread_actions if a["status"] == "Email Unknown"]
            if len(unknown_rows) == 1:
                a = unknown_rows[0]
                update_action_contact(a["row"], responsible_email=email_, responsible_name=name, status="Open")
                append_to_action_list_column(a["row"], 15, email_)
                updated.append(f"Updated '{a['action'][:60]}' — responsible: {name} <{email_}>")
    return updated


def apply_ncr_clarification(ncr_row_number, new_contacts, existing_emails, existing_names):
    emails = list(existing_emails)
    names = list(existing_names)
    added = []
    for c in new_contacts:
        e = c["email"].strip()
        n = c.get("name", "").strip()
        if e.lower() not in [x.lower() for x in emails]:
            emails.append(e); names.append(n)
            added.append(f"{n} <{e}>" if n else e)
    try:
        sheet = get_ncr_tracker_sheet()
        if sheet:
            sheet.update_cell(ncr_row_number, 5, ",".join(emails))
            sheet.update_cell(ncr_row_number, 16, ",".join(names))
            if emails:
                sheet.update_cell(ncr_row_number, 7, "Open")
    except Exception as e:
        logger.error(f"Apply NCR clarification error: {e}")
    return added


def dispatch_approved_external_draft(action_data):
    reminder_count = action_data["reminder_count"] + 1
    draft = draft_external_reminder(action_data["action"], action_data["responsible_name"], action_data["responsible"],
                                    action_data["due_date"], action_data["meeting_ref"], reminder_count,
                                    on_behalf_of=get_first_name(action_data["mom_sender"]))
    if not draft or action_data["email"] in ["UNKNOWN", ""]:
        return
    if not action_data["responsible_name"]:
        domain = action_data["email"].split("@")[-1] if "@" in action_data["email"] else ""
        all_to = get_all_domain_emails(domain, action_data["all_participants"])
        to_emails = all_to if all_to else [action_data["email"]]
    else:
        to_emails = [action_data["email"]]
    cc_list = build_cc_for_external(action_data)
    resp_label = action_data["responsible_name"] or action_data["responsible"]
    reminder_tag = {1: "Follow-up", 2: "Second Follow-up", 3: "Escalation Notice"}.get(reminder_count, "Follow-up")
    html_reminder = build_external_reminder_html(draft, action_data["meeting_ref"], action_data["action"], resp_label, action_data["due_date"], reminder_tag, on_behalf_of=get_first_name(action_data["mom_sender"]))
    sent = send_email(to_emails, f"Action Item Follow-up — {action_data['meeting_ref']}", draft, html_body=html_reminder, cc_emails=cc_list)
    if sent:
        update_action_row(action_data["row"], status="Reminded", last_reminded=datetime.now().strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_count, draft_sent=datetime.now().strftime("%d.%m.%Y %H:%M"))


def handle_mom_thread_reply(sender, body, thread_actions, msg_id_hdr, in_reply_to, references):
    if not thread_actions:
        return False
    meeting_ref = thread_actions[0]["meeting_ref"]
    new_refs = f"{references} {msg_id_hdr}".strip() if references else msg_id_hdr
    cc = [r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()]

    if is_rejection_reply(body):
        for a in thread_actions:
            update_action_row(a["row"], status="Closed — No Action Required", notes=f"Closed per instruction from {sender}")
        notice = f"Dear Team,\n\nPer instruction from {sender}, the following action item(s) from {meeting_ref} have been closed with no further action.\n\n"
        for a in thread_actions:
            notice += f"- {a['action']}\n"
        notice += "\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email(REPORT_RECIPIENTS, f"Closed — No Action Required — {meeting_ref}", notice, html_body=build_reply_html(notice))
        return True

    actions_summary = "\n".join([f"- {a['action']} | Responsible: {a['responsible']} | Email: {a['email']} | Status: {a['status']}" for a in thread_actions])
    clar = extract_mom_clarification(body, actions_summary)
    clarifications = clar.get("clarifications", [])

    if clarifications:
        updated = apply_mom_clarifications(thread_actions, clarifications)
        notice = f"Dear {get_first_name(sender)},\n\nThank you for the clarification.\n\n"
        if updated:
            notice += "I have updated the Action Tracker as follows:\n\n"
            for s in updated:
                notice += f"- {s}\n"
        else:
            notice += "I have noted the information provided.\n\n"
        notice += "\nThe chase protocol will now proceed automatically for these items.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([sender], f"Action Tracker Updated — {meeting_ref}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
        return True

    if is_approval_reply(body):
        for a in thread_actions:
            update_action_row(a["row"], notes="Party identification confirmed by internal team")
        notice = f"Dear {get_first_name(sender)},\n\nThank you for confirming. The chase protocol will proceed automatically per the agreed schedule.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([sender], f"Confirmed — {meeting_ref}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
        return True

    notice = f"Dear {get_first_name(sender)},\n\nThank you for your reply regarding {meeting_ref}. I was unable to identify a specific email address or clear instruction in this message. If you are providing a contact, please state the name or company together with the email address, for example: ABC Contractor — contact@abc.az. If no further action is required please reply no need.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
    send_email([sender], f"Re: {meeting_ref}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
    return True


def handle_ncr_thread_reply(sender, body, ncr_matches, msg_id_hdr, references):
    if not ncr_matches:
        return False
    new_refs = f"{references} {msg_id_hdr}".strip() if references else msg_id_hdr
    all_client_emails = list(set(e for ncr in ncr_matches for e in ncr.get("client_emails", [])))
    cc = list(set([r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()] + all_client_emails))
    ncr_numbers = ", ".join([n["ncr_number"] for n in ncr_matches])

    ncr_summary = "\n".join([f"NCR {n['ncr_number']}: {n['description']} | Contractor: {n['contractor']} | Current contacts: {', '.join(n['all_emails']) or 'None'}" for n in ncr_matches])
    new_contacts = extract_ncr_clarification(body, ncr_summary)

    if new_contacts:
        all_added = []
        for ncr in ncr_matches:
            added = apply_ncr_clarification(ncr["row"], new_contacts, ncr["all_emails"], ncr["all_names"])
            if added:
                all_added.append((ncr["ncr_number"], added))
        notice = f"Dear {get_first_name(sender)},\n\nThank you for the clarification.\n\n"
        if all_added:
            for num, added in all_added:
                notice += f"NCR {num}:\n"
                for a in added:
                    notice += f"- {a}\n"
                notice += "\n"
        else:
            notice += "I have noted the information provided.\n\n"
        notice += "These NCR(s) remain open and follow-up will continue automatically until corrective action reports are received and accepted.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([sender], f"NCR Contacts Updated — {ncr_numbers}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
        return True

    if is_approval_reply(body):
        notice = "Dear " + get_first_name(sender) + ",\n\nThank you for confirming. The following NCR(s) remain open and follow-up will proceed automatically per the chase protocol until corrective action reports are received and accepted:\n\n"
        for ncr in ncr_matches:
            notice += f"- NCR {ncr['ncr_number']}: {ncr['description']}\n"
        notice += "\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([sender], f"Confirmed — {ncr_numbers}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
        return True

    notice = f"Dear {get_first_name(sender)},\n\nThank you for your reply regarding {ncr_numbers}. I was unable to identify a specific contractor contact or clear instruction in this message. If you are providing a contact, please state the name and email address explicitly, for example: John Smith — john@contractor.az.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
    send_email([sender], f"Re: {ncr_numbers}", notice, html_body=build_reply_html(notice), cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=new_refs)
    return True


def handle_daily_report_reply(sender, body, msg_id_hdr, references):
    open_actions, _, _ = read_actions_for_report()
    open_ncrs, _ = read_ncrs_for_report()

    action_sheet = get_action_tracker_sheet()
    ncr_sheet = get_ncr_tracker_sheet()
    action_rows = action_sheet.get_all_values()[1:] if action_sheet else []
    ncr_rows = ncr_sheet.get_all_values()[1:] if ncr_sheet else []

    action_lookup = {}
    for i, row in enumerate(action_rows, start=2):
        if len(row) < 7:
            continue
        status = row[6].strip() if len(row) > 6 else ""
        if status in ["Open", "Reminded", "Draft Pending"]:
            action_lookup[i] = {"row": i, **get_action_data_from_row(row)}

    ncr_lookup = {}
    for i, row in enumerate(ncr_rows, start=2):
        if len(row) < 7:
            continue
        status = row[6].strip() if len(row) > 6 else ""
        if status in ["Open", "Reminded"]:
            ncr_lookup[i] = {"row": i, **get_ncr_data_from_row(row)}

    requests = extract_daily_report_requests(body, list(action_lookup.values()), list(ncr_lookup.values()))

    if not requests:
        notice = f"Dear {get_first_name(sender)},\n\nI could not identify a specific action item or NCR in your reply. Please reference the NCR number or a clear description of the action, for example: send reminder for NCR-0008.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([sender], "Re: SCOPE IQ Daily Report", notice, html_body=build_reply_html(notice))
        return

    summary_lines = []
    cc = [r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()]

    for req in requests:
        ref_type = req.get("ref_type")
        ref = req.get("ref")
        want = req.get("request")

        if ref_type == "ACTION" and ref in action_lookup:
            data = action_lookup[ref]
            if want == "close":
                update_action_row(data["row"], status="Closed — No Action Required", notes=f"Closed per instruction from {sender} via daily report reply")
                summary_lines.append(f"Closed action: {data['action'][:80]}")
            elif want == "reminder" and data["email"] not in ["UNKNOWN", ""]:
                reminder_count = data["reminder_count"] + 1
                draft = draft_external_reminder(data["action"], data["responsible_name"], data["responsible"], data["due_date"], data["meeting_ref"], reminder_count, on_behalf_of=get_first_name(sender))
                if draft:
                    resp_label = data["responsible_name"] or data["responsible"]
                    reminder_tag = {1: "Follow-up", 2: "Second Follow-up", 3: "Escalation Notice"}.get(reminder_count, "Follow-up")
                    to_preview = data["email"]
                    cc_preview = build_cc_for_external(data)
                    approval = f"Dear {get_first_name(sender)},\n\nAs requested, I have prepared a follow-up for the action below.\n\nMeeting reference: {data['meeting_ref']}\nAction: {data['action']}\nResponsible: {resp_label} ({data['email']})\n\nPlease reply with approve or send to dispatch, or no need to cancel.\n\nWhen approved this email will be sent:\nTo: {to_preview}\nCC: {', '.join(cc_preview)}\n\n{'='*50}\nDRAFT — {reminder_tag.upper()} TO {resp_label.upper()}:\n{'='*50}\n\n{draft}\n\n{'='*50}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    sent = send_email([sender], f"Approval Required — {reminder_tag} to {resp_label}", approval, html_body=build_reply_html(approval), cc_emails=cc, reply_to_msg_id=data["thread_id"], references=data["thread_id"])
                    if sent:
                        update_action_row(data["row"], status="Draft Pending", last_reminded=datetime.now().strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_count)
                        summary_lines.append(f"Draft prepared and sent for approval: {data['action'][:80]}")

        elif ref_type == "NCR" and ref in ncr_lookup:
            data = ncr_lookup[ref]
            if want == "close":
                update_ncr_row(data["row"], status="Closed", notes=f"Closed per instruction from {sender} via daily report reply")
                summary_lines.append(f"Closed NCR {data['ncr_number']}")
            elif want == "reminder" and data["all_emails"]:
                reminder_count = data["reminder_count"] + 1
                resp_label = data["responsible_name"] or data["contractor"]
                prompt = f"""Draft a polite and professional reminder email to a contractor regarding an open Non-Conformance Report.

NCR reference: {data['ncr_number']}
Description: {data['description']}
Contractor: {resp_label}

Start with Dear {resp_label if not data['responsible_name'] else data['responsible_name'].split()[0]},
State clearly this NCR remains open and will not be closed until a corrective action report is submitted and accepted.
Write complete formal professional email. No bullet points or symbols.
End with:
Alex Rivera
Construction Expert
SCOPE Consulting MMC
internal@scope-iq.io"""
                response = anthropic_client.messages.create(model=MODEL, max_tokens=700, system=SYSTEM_PROMPT, messages=[{"role": "user", "content": prompt}])
                draft = response.content[0].text
                approval = f"Dear {get_first_name(sender)},\n\nAs requested, I have prepared a follow-up for NCR {data['ncr_number']}.\n\nDescription: {data['description']}\nContractor: {resp_label} ({', '.join(data['all_emails'])})\n\nPlease reply with approve or send to dispatch.\n\n{'='*50}\nDRAFT — NCR FOLLOW-UP TO {resp_label.upper()}:\n{'='*50}\n\n{draft}\n\n{'='*50}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                ncr_reminder_html = build_ncr_reminder_html(approval, data["ncr_number"], data["description"], resp_label, data["date_raised"], 0)
                sent = send_email([sender], f"Approval Required — NCR Follow-up to {resp_label}", approval, html_body=ncr_reminder_html, cc_emails=cc, reply_to_msg_id=data["thread_id"], references=data["thread_id"])
                if sent:
                    update_ncr_row(data["row"], status="Reminded", last_reminded=datetime.now().strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_count)
                    summary_lines.append(f"Draft prepared and sent for approval: NCR {data['ncr_number']}")

    if summary_lines:
        notice = f"Dear {get_first_name(sender)},\n\nFollowing your reply to the Daily Report, I have taken the following action(s):\n\n"
        for line in summary_lines:
            notice += f"- {line}\n"
        notice += "\nWhere a draft was prepared, please check your inbox for the separate approval email before it is sent externally.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
    else:
        notice = f"Dear {get_first_name(sender)},\n\nI identified the item(s) referenced but could not action them — the contact email may be missing or unmatched. Please check the tracker or provide the correct contact.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"

    send_email([sender], "Re: SCOPE IQ Daily Report", notice, html_body=build_reply_html(notice), cc_emails=cc)


def draft_external_reminder(action_item, responsible_name, responsible_party, due_date, meeting_ref, reminder_number, outstanding=None, on_behalf_of=None):
    try:
        tone_map = {1: "polite and professional", 2: "firm and urgent", 3: "formal escalation"}
        if responsible_name and responsible_name.strip():
            salutation = f"Dear {responsible_name.strip().split()[0]},"
        else:
            company = responsible_party.strip().split()[0] if responsible_party else "Team"
            salutation = f"Dear {company} Team,"
        outstanding_text = f"\n\nOutstanding items:\n{outstanding}" if outstanding else ""
        on_behalf_line = f" writing on behalf of {on_behalf_of}, SCOPE Consulting MMC" if on_behalf_of else ""
        prompt = f"""Draft a {tone_map.get(reminder_number, 'formal')} reminder email to an external party.

Salutation: {salutation}
Meeting reference: {meeting_ref}
Action item: {action_item}
Responsible: {responsible_name or responsible_party}
Due date: {due_date}
Reminder number: {reminder_number} of 3{outstanding_text}

Start with exactly: {salutation}
Mention you are{on_behalf_line}, following up on the referenced meeting.
Write complete formal professional email. No bullet points or symbols.
End with:
Alex Rivera
Construction Expert
SCOPE Consulting MMC
internal@scope-iq.io"""
        response = anthropic_client.messages.create(model=MODEL, max_tokens=800, system=SYSTEM_PROMPT, messages=[{"role": "user", "content": prompt}])
        return response.content[0].text
    except Exception as e:
        logger.error(f"Draft reminder error: {e}")
        return None


def get_action_data_from_row(row):
    participants_raw = row[14].strip() if len(row) > 14 else ""
    all_participants = [p.strip() for p in participants_raw.split(",") if p.strip() and "@" in p.strip()]
    client_raw = row[15].strip() if len(row) > 15 else ""
    client_emails = [c.strip() for c in client_raw.split(",") if c.strip() and "@" in c.strip()]
    return {
        "date": row[0].strip() if row[0] else "", "meeting_ref": row[1].strip() if len(row) > 1 else "",
        "action": row[2].strip() if len(row) > 2 else "", "responsible": row[3].strip() if len(row) > 3 else "",
        "email": row[4].strip() if len(row) > 4 else "", "due_date": row[5].strip() if len(row) > 5 else "",
        "status": row[6].strip() if len(row) > 6 else "", "last_reminded": row[7].strip() if len(row) > 7 else "",
        "reminder_count": int(row[8].strip()) if len(row) > 8 and row[8].strip().isdigit() else 0,
        "thread_id": row[9].strip() if len(row) > 9 else "", "mom_sender": row[10].strip() if len(row) > 10 else "",
        "all_participants": all_participants, "client_emails": client_emails,
        "responsible_name": row[16].strip() if len(row) > 16 else ""
    }


def build_cc_for_external(action_data):
    resp_email = action_data["email"].lower()
    cc_list = []
    for addr in action_data["all_participants"]:
        addr_lower = addr.lower()
        if addr_lower == ZOHO_EMAIL.lower() or addr_lower == resp_email:
            continue
        cc_list.append(addr)
    for r in REPORT_RECIPIENTS:
        if r.lower() not in [c.lower() for c in cc_list]:
            cc_list.append(r)
    return list(set(cc_list))


def get_all_domain_emails(domain, all_participants):
    return [p for p in all_participants if "@" in p and p.split("@")[-1].lower() == domain.lower()]


def route_external_reply_for_approval(sender, subject, body, action_data, msg_id_hdr):
    analysis = analyse_external_reply(action_data["action"], body)
    mom_sender = action_data["mom_sender"]
    cc_list = [r for r in REPORT_RECIPIENTS if r.lower() != mom_sender.lower()]
    resp_label = action_data["responsible_name"] or action_data["responsible"]

    if analysis.get("satisfied"):
        update_action_row(action_data["row"], status="Closed", external_reply=body[:300], notes=analysis.get("analysis", ""))
        notice = f"Dear {get_first_name(mom_sender)},\n\nA reply has been received from {resp_label} and I have assessed it as satisfactorily addressing the required action.\n\nMeeting reference: {action_data['meeting_ref']}\nAction: {action_data['action']}\n\nMy assessment:\n{analysis.get('analysis', '')}\n\nThis action has been closed in the Action Tracker.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
        send_email([mom_sender], f"Action Closed — {action_data['action'][:50]}", notice, html_body=build_reply_html(notice), cc_emails=cc_list, reply_to_msg_id=action_data["thread_id"], references=action_data["thread_id"])
    else:
        outstanding = analysis.get("outstanding_items", "")
        draft = draft_external_reminder(action_data["action"], action_data["responsible_name"], action_data["responsible"], "As previously agreed", action_data["meeting_ref"], 1, outstanding, on_behalf_of=get_first_name(mom_sender))
        if draft:
            approval = f"Dear {get_first_name(mom_sender)},\n\nA reply has been received from {resp_label} regarding the action below. My assessment is that it does not fully satisfy the required action.\n\nMeeting reference: {action_data['meeting_ref']}\nAction: {action_data['action']}\n\nMy assessment:\n{analysis.get('analysis', '')}\n\nOutstanding items:\n{outstanding}\n\nI have prepared a follow-up email for your approval. Please reply with approve or send. If no further action is required please reply no need.\n\n{'='*50}\nDRAFT FOLLOW-UP TO {resp_label.upper()}:\n{'='*50}\n\n{draft}\n\n{'='*50}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
            update_action_row(action_data["row"], status="Draft Pending", external_reply=body[:300], notes=f"Reply unsatisfactory: {outstanding[:200]}")
            send_email([mom_sender], f"Approval Required — Follow-up to {resp_label}", approval, html_body=build_reply_html(approval), cc_emails=cc_list, reply_to_msg_id=action_data["thread_id"], references=action_data["thread_id"])


def process_mom_email(sender, subject, body, attachments, all_thread_with_names, msg_id_hdr):
    mom_content = body
    if attachments:
        for att in attachments:
            mom_content += f"\n\n{att['name']}:\n{att['content']}"
    extracted = extract_mom_actions(mom_content, all_thread_with_names, subject)
    meeting_ref = extracted.get("meeting_reference", subject)
    actions = extracted.get("actions", [])
    client_id = extracted.get("client_identified", "Unknown")
    contractor_id = extracted.get("contractor_identified", "Unknown")
    party_classification = extracted.get("party_classification", [])
    internal, clients, others = merge_party_classification(all_thread_with_names, party_classification)
    contractors = others
    client_emails = [p["email"] for p in clients]
    all_participants = [p["email"] for p in all_thread_with_names if p["email"] != ZOHO_EMAIL.lower()]

    if not actions:
        save_to_monitoring(sender, subject, "MOM received — no actions extracted", "Review MOM manually", msg_id_hdr, "Monitoring")
        return

    unknown_count = 0
    action_rows = []
    for action in actions:
        resp_email = action.get("responsible_email", "UNKNOWN")
        resp_name = action.get("responsible_name", "")
        status = "Open" if resp_email != "UNKNOWN" else "Email Unknown"
        if resp_email == "UNKNOWN":
            unknown_count += 1
        save_action_item(meeting_ref, action.get("action", ""), action.get("responsible_party", "Unknown"), resp_email, resp_name, action.get("due_date", "Not specified"), msg_id_hdr, sender, all_participants, client_emails, status)
        action_rows.append({**action, "status": status})

    html_confirmation = build_mom_confirmation_html(sender, meeting_ref, action_rows, client_id, contractor_id, internal, clients, contractors, unknown_count)
    plain = f"MOM Action Items — Confirmation Required — {meeting_ref}\n\nClient identified: {client_id}\nContractor identified: {contractor_id}\n\n"
    for i, action in enumerate(actions, 1):
        plain += f"{i}. {action.get('action','')} — {action.get('responsible_party','Unknown')} ({action.get('responsible_email','UNKNOWN')})\n"
    plain += "\nPlease reply approve/confirmed to proceed, or no need to close this item.\n\nAlex Rivera | SCOPE Consulting MMC"
    cc_approval = [r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()]
    send_email([sender], f"MOM Action Items — Confirmation Required — {meeting_ref}", plain, html_body=html_confirmation, cc_emails=cc_approval, reply_to_msg_id=msg_id_hdr, references=msg_id_hdr)


def process_ncr_email(sender, subject, body, attachments, all_thread_with_names, msg_id_hdr):
    ncr_content = body
    if attachments:
        for att in attachments:
            ncr_content += f"\n\n{att['name']}:\n{att['content']}"
    extracted = extract_ncr_details(ncr_content, all_thread_with_names, subject)
    all_participants = [p["email"] for p in all_thread_with_names if p["email"] != ZOHO_EMAIL.lower()]
    contacts = extracted.get("contacts", [])
    contact_emails = [c["email"] for c in contacts if c.get("email")]
    contact_names = [c.get("name", "") for c in contacts if c.get("email")]
    status = "Open" if contact_emails else "Email Unknown"
    internal, clients, _ = classify_thread_participants(all_thread_with_names)
    client_emails = [p["email"] for p in clients]

    save_ncr_item(extracted.get("ncr_number", "Unknown"), extracted.get("description", subject), extracted.get("contractor", "Unknown"), contact_emails, contact_names, extracted.get("date_raised", "Not specified"), msg_id_hdr, sender, all_participants, client_emails, status)

    contacts_display = "\n".join([f"- {c.get('name') or c['email']} <{c['email']}>" for c in contacts]) if contacts else "No contractor contacts detected."
    client_display = "\n".join([f"- {p['name']} <{p['email']}>" if p['name'] else f"- {p['email']}" for p in clients]) if clients else "No client contacts detected in this thread."

    notification = f"Dear {get_first_name(sender)},\n\nI have logged the following Non-Conformance Report in the NCR Tracker.\n\n"
    notification += f"NCR reference: {extracted.get('ncr_number', 'Unknown')}\nDescription: {extracted.get('description', '')}\nContractor identified: {extracted.get('contractor', 'Unknown')}\nContractor contacts:\n{contacts_display}\nClient contacts (CC for awareness):\n{client_display}\nDate raised: {extracted.get('date_raised', 'Not specified')}\n\n"
    if not contact_emails:
        notification += "I was unable to match the contractor to a contact email. Please provide the correct contact(s).\n\n"
    if not clients:
        notification += "I was also unable to identify a client contact for this NCR. Please provide the relevant client team member(s).\n\n"
    notification += "All SCOPE team members and identified client contacts will be CC'd on every follow-up. This NCR will remain open until a corrective action report is received and accepted. Please reply confirmed if correct, or provide additional or corrected contacts.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"

    cc = list(set([r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()] + client_emails))
    ncr_html = build_ncr_confirmation_html(sender, {**extracted, "contractor_email": ", ".join(contact_emails) if contact_emails else "UNKNOWN", "contractor_name": ", ".join([n for n in contact_names if n])})
    send_email([sender], f"NCR Logged — {extracted.get('ncr_number', 'Unknown')}", notification, html_body=ncr_html, cc_emails=cc, reply_to_msg_id=msg_id_hdr, references=msg_id_hdr)


def build_mom_confirmation_html(sender, meeting_ref, actions, client_id, contractor_id, internal, clients, contractors, unknown_count):
    today = datetime.now().strftime("%d %B %Y")
    time_now = datetime.now().strftime("%H:%M")

    def fmt_party_card(title, lst, note, empty_note, accent):
        if lst:
            rows = "".join([f'<div style="font-size:12px;color:#333;padding:3px 0;">{p["name"]} &lt;{p["email"]}&gt;</div>' if p["name"] else f'<div style="font-size:12px;color:#333;padding:3px 0;">{p["email"]}</div>' for p in lst])
            body = rows + f'<div style="font-size:11px;color:#888;margin-top:6px;">{note}</div>'
        else:
            body = f'<div style="font-size:12px;color:#888;">{empty_note}</div>'
        return f'<div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:12px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 14px;border-left:3px solid {accent};"><span style="font-size:11px;color:#555;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">{title}</span></div><div style="padding:12px 14px;">{body}</div></div>'

    party_html = fmt_party_card(f"Client — {client_id}", clients, "CC'd for awareness on all follow-ups.", "No client email detected in this thread.", "#3CB496")
    roles_present = {}
    for p in contractors:
        roles_present.setdefault(p.get("role_label", "Contractor"), []).append(p)
    if roles_present:
        for role, people in roles_present.items():
            party_html += fmt_party_card(f"{role} — {contractor_id if role.lower() == 'contractor' else ''}".strip(" —"), people, "Will receive action follow-up emails once you approve.", f"No {role.lower()} email detected.", "#f0a030")
    else:
        party_html += fmt_party_card(f"Contractor — {contractor_id}", [], "Will receive action follow-up emails once you approve.", "No contractor email detected.", "#f0a030")
    party_html += fmt_party_card("SCOPE team", internal, "CC'd on all outgoing correspondence.", "No SCOPE team members detected in thread.", "#1a2942")

    actions_html = ""
    for i, a in enumerate(actions, 1):
        role = a.get("responsible_role_label", "Contractor")
        name = a.get("responsible_name", "")
        party = a.get("responsible_party", "Unknown")
        label = f"{name} ({party})" if name else party
        email_ = a.get("responsible_email", "UNKNOWN")
        due = a.get("due_date", "Not specified")
        status = a.get("status", "Open")
        role_bg = {"Client": "#e1f5ee", "SCOPE": "#eef1f6"}.get(role, "#fff8ee")
        role_tx = {"Client": "#0f6e56", "SCOPE": "#1a2942"}.get(role, "#9a6000")
        status_bg = "#fff0f0" if status == "Email Unknown" else "#fff8ee"
        status_tx = "#c00000" if status == "Email Unknown" else "#9a6000"
        actions_html += f'<div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:12px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 14px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;">Action {i} of {len(actions)}</span><div><span style="background:{role_bg};color:{role_tx};font-size:10px;padding:2px 8px;border-radius:20px;margin-right:6px;">{role}</span><span style="background:{status_bg};color:{status_tx};font-size:10px;padding:2px 8px;border-radius:20px;">{status}</span></div></div><div style="padding:14px;"><div style="font-size:13px;color:#1a2942;font-weight:600;margin-bottom:8px;">{a.get("action","")}</div><table style="width:100%;font-size:12px;border-collapse:collapse;"><tr><td style="color:#888;padding:3px 0;width:120px;">Responsible</td><td style="color:#333;">{label}</td></tr><tr><td style="color:#888;padding:3px 0;">Email</td><td style="color:#333;">{email_}</td></tr><tr><td style="color:#888;padding:3px 0;">Due date</td><td style="color:#333;">{due}</td></tr></table></div></div>'

    unknown_banner = f'<div style="background:#fff0f0;border:1px solid #f0c0c0;border-radius:6px;padding:10px 14px;margin-bottom:16px;"><div style="font-size:11px;font-weight:600;color:#c00000;margin-bottom:4px;">EMAIL UNKNOWN — ACTION REQUIRED</div><div style="font-size:12px;color:#800000;">{unknown_count} action(s) could not be matched to a contact email.</div></div>' if unknown_count else ""

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:24px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="color:#fff;font-size:22px;font-weight:600;letter-spacing:1px;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="background:#3CB496;color:#fff;font-size:11px;padding:4px 12px;border-radius:20px;font-weight:500;">MOM Action Items</div>
    </div>
    <table style="width:100%;font-size:12px;border-collapse:collapse;">
      <tr><td style="color:#8facc8;padding:2px 0;width:50%;">Date &nbsp;<strong style="color:#c8ddf0;">{today}</strong></td><td style="color:#8facc8;padding:2px 0;">Time &nbsp;<strong style="color:#c8ddf0;">{time_now} Baku</strong></td></tr>
      <tr><td style="color:#8facc8;padding:2px 0;">Prepared by &nbsp;<strong style="color:#c8ddf0;">Alex Rivera</strong></td><td style="color:#8facc8;padding:2px 0;">Meeting ref &nbsp;<strong style="color:#c8ddf0;">{meeting_ref}</strong></td></tr>
    </table>
  </div>
  <div style="background:#243550;padding:12px 28px;"><div style="font-size:11px;color:#8facc8;text-transform:uppercase;letter-spacing:0.5px;">Total actions identified</div><div style="font-size:22px;color:#fff;font-weight:600;">{len(actions)}</div></div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:24px 28px;">
    <p style="font-size:14px;color:#444;line-height:1.7;margin:0 0 16px;">Dear {get_first_name(sender)}, I have analysed the Minutes of Meeting and extracted the action items below.</p>
    <div style="font-size:11px;font-weight:600;color:#888;letter-spacing:1px;text-transform:uppercase;margin-bottom:10px;">Party identification — please confirm</div>
    {party_html}
    {unknown_banner}
    <div style="font-size:11px;font-weight:600;color:#888;letter-spacing:1px;text-transform:uppercase;margin:20px 0 10px;">Extracted action items</div>
    {actions_html}
    <div style="background:#f0f7ff;border:1px solid #cfe3fb;border-radius:6px;padding:12px 14px;margin-top:16px;"><div style="font-size:12px;color:#1a4d80;line-height:1.6;">Please reply <strong>approve</strong> or <strong>confirmed</strong> if correct. If no action is required, please reply <strong>no need</strong>. If any party is incorrect, state the correct name/email. Please reply directly to this email.</div></div>
    <div style="border-top:1px solid #f0f0f0;margin-top:20px;padding-top:16px;"><div style="font-size:13px;font-weight:600;color:#1a2942;">Alex Rivera</div><div style="font-size:12px;color:#666;">Construction Expert</div><div style="font-size:12px;color:#666;">SCOPE Consulting MMC</div><div style="font-size:12px;color:#3CB496;">internal@scope-iq.io</div></div>
  </div>
</div></body></html>"""


def build_ncr_confirmation_html(sender, ncr_data):
    today = datetime.now().strftime("%d %B %Y")
    time_now = datetime.now().strftime("%H:%M")
    ncr_number = ncr_data.get("ncr_number", "Unknown")
    description = ncr_data.get("description", "")
    contractor = ncr_data.get("contractor", "Unknown")
    contractor_name = ncr_data.get("contractor_name", "")
    email_ = ncr_data.get("contractor_email", "UNKNOWN")
    date_raised = ncr_data.get("date_raised", "Not specified")
    email_unknown = email_ == "UNKNOWN"
    status_bg = "#fff0f0" if email_unknown else "#fff8ee"
    status_tx = "#c00000" if email_unknown else "#9a6000"
    status_label = "Email Unknown" if email_unknown else "Open"
    contact_line = f"{contractor_name} &lt;{email_}&gt;" if contractor_name and not email_unknown else (email_ if not email_unknown else "No contact email matched")
    unknown_banner = '<div style="background:#fff0f0;border:1px solid #f0c0c0;border-radius:6px;padding:10px 14px;margin-bottom:16px;"><div style="font-size:11px;font-weight:600;color:#c00000;margin-bottom:4px;">EMAIL UNKNOWN — ACTION REQUIRED</div><div style="font-size:12px;color:#800000;">Could not match the contractor to a contact email.</div></div>' if email_unknown else ""

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:24px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="color:#fff;font-size:22px;font-weight:600;letter-spacing:1px;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="background:#c00000;color:#fff;font-size:11px;padding:4px 12px;border-radius:20px;font-weight:500;">NCR Logged</div>
    </div>
    <table style="width:100%;font-size:12px;border-collapse:collapse;">
      <tr><td style="color:#8facc8;padding:2px 0;width:50%;">Date &nbsp;<strong style="color:#c8ddf0;">{today}</strong></td><td style="color:#8facc8;padding:2px 0;">Time &nbsp;<strong style="color:#c8ddf0;">{time_now} Baku</strong></td></tr>
      <tr><td style="color:#8facc8;padding:2px 0;">Prepared by &nbsp;<strong style="color:#c8ddf0;">Alex Rivera</strong></td><td style="color:#8facc8;padding:2px 0;">NCR ref &nbsp;<strong style="color:#c8ddf0;">{ncr_number}</strong></td></tr>
    </table>
  </div>
  <div style="background:#3a1a1a;padding:16px 28px;"><div style="font-size:11px;color:#e0a8a8;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">Non-conformance description</div><div style="font-size:14px;color:#fff;font-weight:600;line-height:1.5;">{description}</div></div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:24px 28px;">
    <p style="font-size:14px;color:#444;line-height:1.7;margin:0 0 16px;">Dear {get_first_name(sender)}, I have logged the following Non-Conformance Report in the NCR Tracker.</p>
    <div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:16px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 14px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;font-weight:500;">NCR DETAILS</span><span style="background:{status_bg};color:{status_tx};font-size:10px;padding:2px 8px;border-radius:20px;">{status_label}</span></div><div style="padding:14px 16px;"><table style="width:100%;font-size:13px;border-collapse:collapse;"><tr><td style="color:#888;padding:4px 0;width:140px;">NCR reference</td><td style="color:#1a2942;font-weight:600;">{ncr_number}</td></tr><tr><td style="color:#888;padding:4px 0;">Contractor</td><td style="color:#333;">{contractor}</td></tr><tr><td style="color:#888;padding:4px 0;">Contact</td><td style="color:#333;">{contact_line}</td></tr><tr><td style="color:#888;padding:4px 0;">Date raised</td><td style="color:#333;">{date_raised}</td></tr></table></div></div>
    {unknown_banner}
    <div style="background:#fff8ee;border:1px solid #f0c060;border-radius:6px;padding:12px 14px;margin-top:8px;"><div style="font-size:12px;color:#5a3a00;line-height:1.6;">This NCR will remain <strong>open</strong> until a corrective action report is received and accepted.</div></div>
    <div style="background:#f0f7ff;border:1px solid #cfe3fb;border-radius:6px;padding:12px 14px;margin-top:12px;"><div style="font-size:12px;color:#1a4d80;line-height:1.6;">Please reply <strong>confirmed</strong> if correct, or provide corrections/additional contacts.</div></div>
    <div style="border-top:1px solid #f0f0f0;margin-top:20px;padding-top:16px;"><div style="font-size:13px;font-weight:600;color:#1a2942;">Alex Rivera</div><div style="font-size:12px;color:#666;">Construction Expert</div><div style="font-size:12px;color:#666;">SCOPE Consulting MMC</div><div style="font-size:12px;color:#3CB496;">internal@scope-iq.io</div></div>
  </div>
</div></body></html>"""


def build_external_reminder_html(body_text, meeting_ref, action_item, responsible_label, due_date, reminder_label="Follow-up", on_behalf_of=None):
    today = datetime.now().strftime("%d %B %Y")
    time_now = datetime.now().strftime("%H:%M")
    paragraphs = body_text.strip().split("\n\n")
    html_body = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if "Alex Rivera" in para and "SCOPE Consulting" in para:
            lines = para.split("\n")
            sig_html = ""
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "Alex Rivera" in line:
                    sig_html += f'<div style="font-size:13px;font-weight:600;color:#1a2942;">{line}</div>'
                elif "internal@scope-iq.io" in line:
                    sig_html += f'<div style="font-size:12px;color:#3CB496;">{line}</div>'
                else:
                    sig_html += f'<div style="font-size:12px;color:#666;">{line}</div>'
            html_body += f'<div style="margin-top:20px;padding-top:16px;border-top:1px solid #f0f0f0;line-height:1.8;">{sig_html}</div>'
        else:
            lines = para.split("\n")
            para_html = "<br>".join(line.strip() for line in lines if line.strip())
            html_body += f'<p style="font-size:14px;color:#333;line-height:1.8;margin:0 0 16px;">{para_html}</p>'
    badge_color = "#f0a030" if reminder_label == "Follow-up" else "#e07030" if reminder_label == "Second Follow-up" else "#c00000"
    from_label = f"Alex Rivera on behalf of {on_behalf_of}" if on_behalf_of else "Alex Rivera"
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:24px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="color:#fff;font-size:22px;font-weight:600;letter-spacing:1px;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="background:{badge_color};color:#fff;font-size:11px;padding:4px 12px;border-radius:20px;font-weight:500;">{reminder_label}</div>
    </div>
    <table style="width:100%;font-size:12px;border-collapse:collapse;">
      <tr><td style="color:#8facc8;padding:2px 0;width:50%;">Date &nbsp;<strong style="color:#c8ddf0;">{today}</strong></td><td style="color:#8facc8;padding:2px 0;">Time &nbsp;<strong style="color:#c8ddf0;">{time_now} Baku</strong></td></tr>
      <tr><td style="color:#8facc8;padding:2px 0;">From &nbsp;<strong style="color:#c8ddf0;">{from_label}</strong></td><td style="color:#8facc8;padding:2px 0;">Meeting ref &nbsp;<strong style="color:#c8ddf0;">{meeting_ref}</strong></td></tr>
    </table>
  </div>
  <div style="background:#243550;padding:16px 28px;"><div style="font-size:11px;color:#8facc8;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">Action item under follow-up</div><div style="font-size:14px;color:#fff;font-weight:600;line-height:1.5;">{action_item}</div></div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:24px 28px;">
    <div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:20px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 16px;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;font-weight:500;">ACTION DETAILS</span></div><div style="padding:14px 16px;"><table style="width:100%;font-size:13px;border-collapse:collapse;"><tr><td style="color:#888;padding:4px 0;width:130px;">Responsible</td><td style="color:#1a2942;font-weight:600;">{responsible_label}</td></tr><tr><td style="color:#888;padding:4px 0;">Due date</td><td style="color:#333;">{due_date}</td></tr><tr><td style="color:#888;padding:4px 0;">Meeting reference</td><td style="color:#333;">{meeting_ref}</td></tr></table></div></div>
    {html_body}
    <div style="background:#f8f9fa;border-radius:6px;padding:10px 14px;margin-top:16px;"><div style="font-size:11px;color:#888;line-height:1.6;">This follow-up was prepared by <strong style="color:#1a2942;">{from_label}</strong>, using <strong style="color:#3CB496;">SCOPE IQ</strong>.</div></div>
  </div>
</div></body></html>"""


def build_ncr_reminder_html(body_text, ncr_number, description, contractor_label, date_raised, days_open):
    today = datetime.now().strftime("%d %B %Y")
    time_now = datetime.now().strftime("%H:%M")
    paragraphs = body_text.strip().split("\n\n")
    html_body = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if "Alex Rivera" in para and "SCOPE Consulting" in para:
            lines = para.split("\n")
            sig_html = ""
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "Alex Rivera" in line:
                    sig_html += f'<div style="font-size:13px;font-weight:600;color:#1a2942;">{line}</div>'
                elif "internal@scope-iq.io" in line:
                    sig_html += f'<div style="font-size:12px;color:#3CB496;">{line}</div>'
                else:
                    sig_html += f'<div style="font-size:12px;color:#666;">{line}</div>'
            html_body += f'<div style="margin-top:20px;padding-top:16px;border-top:1px solid #f0f0f0;line-height:1.8;">{sig_html}</div>'
        else:
            lines = para.split("\n")
            para_html = "<br>".join(line.strip() for line in lines if line.strip())
            html_body += f'<p style="font-size:14px;color:#333;line-height:1.8;margin:0 0 16px;">{para_html}</p>'
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:24px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="color:#fff;font-size:22px;font-weight:600;letter-spacing:1px;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="background:#c00000;color:#fff;font-size:11px;padding:4px 12px;border-radius:20px;font-weight:500;">NCR Follow-up</div>
    </div>
    <table style="width:100%;font-size:12px;border-collapse:collapse;">
      <tr><td style="color:#8facc8;padding:2px 0;width:50%;">Date &nbsp;<strong style="color:#c8ddf0;">{today}</strong></td><td style="color:#8facc8;padding:2px 0;">Time &nbsp;<strong style="color:#c8ddf0;">{time_now} Baku</strong></td></tr>
      <tr><td style="color:#8facc8;padding:2px 0;">Days open &nbsp;<strong style="color:#c8ddf0;">{days_open}</strong></td><td style="color:#8facc8;padding:2px 0;">NCR ref &nbsp;<strong style="color:#c8ddf0;">{ncr_number}</strong></td></tr>
    </table>
  </div>
  <div style="background:#3a1a1a;padding:16px 28px;"><div style="font-size:11px;color:#e0a8a8;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">Non-conformance under follow-up</div><div style="font-size:14px;color:#fff;font-weight:600;line-height:1.5;">{description}</div></div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:24px 28px;">
    <div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:20px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 16px;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;font-weight:500;">NCR DETAILS</span></div><div style="padding:14px 16px;"><table style="width:100%;font-size:13px;border-collapse:collapse;"><tr><td style="color:#888;padding:4px 0;width:140px;">Contractor</td><td style="color:#1a2942;font-weight:600;">{contractor_label}</td></tr><tr><td style="color:#888;padding:4px 0;">Date raised</td><td style="color:#333;">{date_raised}</td></tr><tr><td style="color:#888;padding:4px 0;">NCR reference</td><td style="color:#333;">{ncr_number}</td></tr></table></div></div>
    {html_body}
    <div style="background:#fff8ee;border:1px solid #f0c060;border-radius:6px;padding:10px 14px;margin-top:12px;"><div style="font-size:12px;color:#5a3a00;line-height:1.6;">This NCR remains open until a corrective action report is received and accepted.</div></div>
    <div style="background:#f8f9fa;border-radius:6px;padding:10px 14px;margin-top:16px;"><div style="font-size:11px;color:#888;line-height:1.6;">This follow-up was prepared by <strong style="color:#1a2942;">Alex Rivera</strong>, using <strong style="color:#3CB496;">SCOPE IQ</strong>.</div></div>
  </div>
</div></body></html>"""


def build_report_html(pending, closed, today, day_name, time_now, open_actions=None, flagged=None, open_ncrs=None):
    n_open = len([r for r in pending if r.get("Status") == "Open"])
    n_monitor = len([r for r in pending if r.get("Status") == "Monitoring"])
    n_closed = len(closed)
    open_actions = open_actions or []
    flagged = flagged or []
    open_ncrs = open_ncrs or []
    total_open = len(pending) + len(open_actions) + len(open_ncrs)

    items_html = ""
    for i, r in enumerate(pending[-15:], 1):
        subj = r.get("Subject") or "No subject"
        sender = r.get("Sender") or "Unknown"
        date = r.get("Date") or "Not recorded"
        status = r.get("Status") or "Open"
        summary = r.get("Summary") or "No summary"
        action = r.get("Action") or "Review required"
        sbg = "#fff8ee" if status == "Open" else "#e1f5ee"
        stx = "#9a6000" if status == "Open" else "#0f6e56"
        items_html += f'<div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:16px;overflow:hidden;"><div style="background:#f8f9fa;padding:12px 16px;display:flex;justify-content:space-between;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;">ITEM {i} OF {len(pending)}</span><span style="background:{sbg};color:{stx};font-size:11px;padding:3px 10px;border-radius:20px;">{status}</span></div><div style="padding:16px;"><table style="width:100%;font-size:13px;border-collapse:collapse;"><tr><td style="color:#888;padding:4px 0;width:120px;">Subject</td><td style="color:#1a2942;font-weight:600;">{subj}</td></tr><tr><td style="color:#888;padding:4px 0;">From</td><td style="color:#333;">{sender}</td></tr><tr><td style="color:#888;padding:4px 0;">Date</td><td style="color:#333;">{date}</td></tr></table><div style="border-top:1px solid #f0f0f0;margin:12px 0;"></div><div style="font-size:11px;color:#888;text-transform:uppercase;margin-bottom:4px;">Summary</div><div style="font-size:13px;color:#444;line-height:1.6;margin-bottom:10px;">{summary}</div><div style="background:#fff8ee;border:1px solid #f0c060;border-radius:6px;padding:10px 12px;"><div style="font-size:10px;font-weight:600;color:#9a6000;text-transform:uppercase;margin-bottom:4px;">Action required</div><div style="font-size:13px;color:#5a3a00;">{action}</div></div></div></div>'

    ext_html = ""
    if open_actions:
        ext_html += '<div style="margin-top:24px;border-top:2px solid #e8e8e8;padding-top:20px;"><div style="font-size:11px;font-weight:600;color:#888;letter-spacing:1px;text-transform:uppercase;margin-bottom:12px;">External action tracker</div>'
        for i, r in enumerate(open_actions[-10:], 1):
            a = r.get("Action Item","") or "No description"
            p = r.get("Responsible Party","") or "Unknown"
            n = r.get("Responsible Name","") or ""
            e = r.get("Responsible Email","") or "Unknown"
            d = r.get("Due Date","") or "Not specified"
            s = r.get("Status","") or "Open"
            m = r.get("Meeting Reference","") or "Unknown"
            display = f"{n} ({p})" if n else p
            ext_html += f'<div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:12px;overflow:hidden;"><div style="background:#f8f9fa;padding:10px 14px;display:flex;justify-content:space-between;border-bottom:1px solid #e8e8e8;"><span style="font-size:12px;color:#888;">External Action {i}</span><span style="background:#fff8ee;color:#9a6000;font-size:11px;padding:2px 8px;border-radius:20px;">{s}</span></div><div style="padding:14px;"><div style="font-size:13px;color:#1a2942;font-weight:600;margin-bottom:8px;">{a}</div><table style="width:100%;font-size:12px;border-collapse:collapse;"><tr><td style="color:#888;padding:3px 0;width:120px;">Responsible</td><td style="color:#333;">{display}</td></tr><tr><td style="color:#888;padding:3px 0;">Email</td><td style="color:#333;">{e}</td></tr><tr><td style="color:#888;padding:3px 0;">Due date</td><td style="color:#333;">{d}</td></tr><tr><td style="color:#888;padding:3px 0;">Meeting</td><td style="color:#333;">{m}</td></tr></table></div></div>'
        if flagged:
            ext_html += f'<div style="background:#fff0f0;border:1px solid #f0c0c0;border-radius:6px;padding:10px 14px;"><div style="font-size:11px;font-weight:600;color:#c00000;margin-bottom:4px;">EMAIL UNKNOWN</div><div style="font-size:12px;color:#800000;">{len(flagged)} action(s) unmatched.</div></div>'
        ext_html += "</div>"

    ncr_html = ""
    if open_ncrs:
        ncr_html += '<div style="margin-top:24px;border-top:2px solid #e8e8e8;padding-top:20px;"><div style="font-size:11px;font-weight:600;color:#888;letter-spacing:1px;text-transform:uppercase;margin-bottom:12px;">Open non-conformance reports</div>'
        for i, r in enumerate(open_ncrs[-10:], 1):
            num = r.get("NCR Number","") or "Unknown"
            desc = r.get("Description","") or "No description"
            contr = r.get("Contractor","") or "Unknown"
            resname = r.get("Responsible Name","") or ""
            e = r.get("Contractor Email","") or "Unknown"
            dr = r.get("Date Raised","") or "Not specified"
            s = r.get("Status","") or "Open"
            display = f"{resname} ({contr})" if resname else contr
            status_bg = "#fff0f0" if s == "Email Unknown" else "#fff8ee"
            status_tx = "#c00000" if s == "Email Unknown" else "#9a6000"
            ncr_html += f'<div style="border:1px solid #e8e8e8;border-radius:8px;margin-bottom:12px;overflow:hidden;"><div style="background:#3a1a1a;padding:10px 14px;display:flex;justify-content:space-between;align-items:center;"><span style="font-size:12px;color:#e0a8a8;">NCR {num}</span><span style="background:{status_bg};color:{status_tx};font-size:11px;padding:2px 8px;border-radius:20px;">{s}</span></div><div style="padding:14px;"><div style="font-size:13px;color:#1a2942;font-weight:600;margin-bottom:8px;">{desc}</div><table style="width:100%;font-size:12px;border-collapse:collapse;"><tr><td style="color:#888;padding:3px 0;width:120px;">Contractor</td><td style="color:#333;">{display}</td></tr><tr><td style="color:#888;padding:3px 0;">Email</td><td style="color:#333;">{e}</td></tr><tr><td style="color:#888;padding:3px 0;">Date raised</td><td style="color:#333;">{dr}</td></tr></table></div></div>'
        ncr_html += '<div style="background:#fff8ee;border:1px solid #f0c060;border-radius:6px;padding:10px 14px;margin-top:8px;"><div style="font-size:11px;color:#5a3a00;">NCRs remain open until a corrective action report is received and accepted.</div></div>'
        ncr_html += "</div>"

    no_items = ""
    if not pending and not open_actions and not open_ncrs:
        no_items = '<div style="text-align:center;padding:32px;"><div style="width:48px;height:48px;border-radius:50%;background:#e1f5ee;margin:0 auto 12px;font-size:22px;color:#3CB496;display:flex;align-items:center;justify-content:center;">&#10003;</div><div style="font-size:15px;color:#333;font-weight:500;">All clear</div><div style="font-size:13px;color:#888;margin-top:4px;">No outstanding items as of today.</div></div>'

    greeting = f"Good morning. Daily report for <strong>{today}</strong>. <strong>{total_open} open item(s)</strong> require attention." if total_open else f"Good morning. Daily report for <strong>{today}</strong>. All items are clear."

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:24px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="color:#fff;font-size:22px;font-weight:600;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="background:#3CB496;color:#fff;font-size:11px;padding:4px 12px;border-radius:20px;">Daily Report</div>
    </div>
    <table style="width:100%;font-size:12px;border-collapse:collapse;">
      <tr><td style="color:#8facc8;padding:2px 0;width:50%;">Date &nbsp;<strong style="color:#c8ddf0;">{day_name}, {today}</strong></td><td style="color:#8facc8;padding:2px 0;">Time &nbsp;<strong style="color:#c8ddf0;">{time_now} Baku</strong></td></tr>
      <tr><td style="color:#8facc8;padding:2px 0;">Prepared by &nbsp;<strong style="color:#c8ddf0;">Alex Rivera</strong></td><td style="color:#8facc8;padding:2px 0;">Status &nbsp;<strong style="color:#c8ddf0;">{"All clear" if total_open == 0 else f"{total_open} open"}</strong></td></tr>
    </table>
  </div>
  <div style="background:#243550;padding:12px 28px;display:flex;">
    <div style="flex:1;text-align:center;border-right:1px solid #1a2942;"><div style="font-size:20px;font-weight:600;color:#f0a030;">{n_open}</div><div style="font-size:10px;color:#8facc8;margin-top:2px;">Internal Open</div></div>
    <div style="flex:1;text-align:center;border-right:1px solid #1a2942;"><div style="font-size:20px;font-weight:600;color:#3CB496;">{n_monitor}</div><div style="font-size:10px;color:#8facc8;margin-top:2px;">Monitoring</div></div>
    <div style="flex:1;text-align:center;border-right:1px solid #1a2942;"><div style="font-size:20px;font-weight:600;color:#e07030;">{len(open_actions)}</div><div style="font-size:10px;color:#8facc8;margin-top:2px;">External Actions</div></div>
    <div style="flex:1;text-align:center;border-right:1px solid #1a2942;"><div style="font-size:20px;font-weight:600;color:#e05050;">{len(open_ncrs)}</div><div style="font-size:10px;color:#8facc8;margin-top:2px;">Open NCRs</div></div>
    <div style="flex:1;text-align:center;"><div style="font-size:20px;font-weight:600;color:#6ab87a;">{n_closed}</div><div style="font-size:10px;color:#8facc8;margin-top:2px;">Closed</div></div>
  </div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:24px 28px;">
    <p style="font-size:14px;color:#444;line-height:1.7;margin:0 0 20px;">{greeting}</p>
    {"<div style='font-size:11px;font-weight:600;color:#888;letter-spacing:1px;text-transform:uppercase;margin-bottom:12px;'>Internal outstanding items</div>" if pending else ""}
    {items_html}
    {ext_html}
    {ncr_html}
    {no_items}
    <div style="border-top:1px solid #f0f0f0;margin-top:24px;padding-top:20px;display:flex;justify-content:space-between;">
      <div style="font-size:12px;color:#666;line-height:1.8;"><strong style="color:#1a2942;font-size:13px;">Alex Rivera</strong><br>Construction Expert<br>SCOPE Consulting MMC<br><span style="color:#3CB496;">internal@scope-iq.io</span></div>
      <div style="font-size:11px;color:#aaa;text-align:right;line-height:1.7;">Generated automatically<br>SCOPE IQ<br>09:00 Baku daily</div>
    </div>
    <div style="background:#f8f9fa;border-radius:6px;padding:10px 14px;margin-top:16px;"><div style="font-size:11px;color:#888;"><strong style="color:#555;">Chase protocol:</strong> Draft at day 3, 7, 14 &nbsp;·&nbsp; Auto-close at day 21 (NCRs remain open until resolved). Reply to this report anytime with an instruction such as send reminder for NCR-0008 to trigger a follow-up on demand.</div></div>
  </div>
</div></body></html>"""


def check_mom_confirmation_and_rejections():
    mail = None
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return
        all_values = sheet.get_all_values()
        if not all_values or len(all_values) < 2:
            return
        open_thread_ids = set()
        for row in all_values[1:]:
            if len(row) < 10:
                continue
            status = row[6].strip() if len(row) > 6 else ""
            thread_id = row[9].strip() if len(row) > 9 else ""
            if thread_id and status not in ["Closed", "Closed — No Response", "Closed — No Action Required"]:
                open_thread_ids.add(thread_id)
        if not open_thread_ids:
            return
        mail = get_imap_connection()
        mail.select("INBOX")
        typ, data = mail.search(None, "ALL")
        if typ != "OK" or not data or not data[0]:
            safe_logout(mail)
            return
        for eid in data[0].split()[-100:]:
            try:
                typ, msg_data = mail.fetch(eid, "(RFC822)")
                if typ != "OK" or not msg_data or not msg_data[0]:
                    continue
                msg = email.message_from_bytes(msg_data[0][1])
                from_addr = extract_email_address(msg.get("From", ""))
                in_reply_to = safe_decode(msg.get("In-Reply-To", "")).strip()
                references = safe_decode(msg.get("References", "")).strip()
                if not is_internal_email(from_addr):
                    continue
                body = strip_quoted_reply(get_email_body(msg))
                if not is_rejection_reply(body):
                    continue
                for thread_id in list(open_thread_ids):
                    if thread_id in in_reply_to or thread_id in references:
                        matches = find_all_actions_by_thread(thread_id)
                        for m in matches:
                            update_action_row(m["row"], status="Closed — No Action Required", notes=f"Closed per instruction from {from_addr}")
                        if matches:
                            open_thread_ids.discard(thread_id)
                            first_action = matches[0]
                            notice = f"Dear Team,\n\nPer your instruction, the following action item(s) from {first_action['meeting_ref']} have been closed with no further action.\n\n"
                            for m in matches:
                                notice += f"- {m['action']}\n"
                            notice += "\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                            send_email(REPORT_RECIPIENTS, f"Closed — No Action Required — {first_action['meeting_ref']}", notice, html_body=build_reply_html(notice))
            except (imaplib.IMAP4.abort, OSError, EOFError):
                safe_logout(mail)
                try:
                    mail = get_imap_connection(); mail.select("INBOX")
                except Exception:
                    break
                continue
            except Exception as e:
                logger.error(f"MOM rejection check error: {e}")
                continue
        safe_logout(mail)
    except Exception as e:
        logger.error(f"MOM confirmation/rejection check error: {e}")
        if mail:
            safe_logout(mail)


def check_external_action_replies():
    mail = None
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return
        all_values = sheet.get_all_values()
        if not all_values or len(all_values) < 2:
            return
        open_actions = {}
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) < 7:
                continue
            data = get_action_data_from_row(row)
            if data["status"] in ["Open", "Reminded", "Draft Sent"] and data["thread_id"] and data["email"]:
                open_actions[data["thread_id"]] = {"row": i, **data}
        if not open_actions:
            return
        mail = get_imap_connection()
        mail.select("INBOX")
        typ, data = mail.search(None, "ALL")
        if typ != "OK" or not data or not data[0]:
            safe_logout(mail)
            return
        for eid in data[0].split()[-100:]:
            try:
                typ, msg_data = mail.fetch(eid, "(RFC822)")
                if typ != "OK" or not msg_data or not msg_data[0]:
                    continue
                msg = email.message_from_bytes(msg_data[0][1])
                in_reply_to = safe_decode(msg.get("In-Reply-To", "")).strip()
                references = safe_decode(msg.get("References", "")).strip()
                from_addr = extract_email_address(msg.get("From", ""))
                if ZOHO_EMAIL.lower() in from_addr or is_internal_email(from_addr):
                    continue
                for thread_id, action_data in open_actions.items():
                    if thread_id and (thread_id in in_reply_to or thread_id in references):
                        reply_body = get_email_body(msg)
                        route_external_reply_for_approval(from_addr, safe_decode(msg.get("Subject", "")), reply_body, action_data, safe_decode(msg.get("Message-ID", "")).strip())
            except (imaplib.IMAP4.abort, OSError, EOFError):
                safe_logout(mail)
                try:
                    mail = get_imap_connection(); mail.select("INBOX")
                except Exception:
                    break
                continue
            except Exception as e:
                logger.error(f"Reply check error: {e}")
                continue
        safe_logout(mail)
    except Exception as e:
        logger.error(f"External reply check error: {e}")
        if mail:
            safe_logout(mail)


def check_action_approvals():
    mail = None
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return
        all_values = sheet.get_all_values()
        pending_approvals = {}
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) < 7:
                continue
            data = get_action_data_from_row(row)
            if data["status"] == "Draft Pending" and data["thread_id"]:
                pending_approvals.setdefault(data["thread_id"], []).append({"row": i, **data})
        if not pending_approvals:
            return
        mail = get_imap_connection()
        mail.select("INBOX")
        typ, data = mail.search(None, "ALL")
        if typ != "OK" or not data or not data[0]:
            safe_logout(mail)
            return
        for eid in data[0].split()[-50:]:
            try:
                typ, msg_data = mail.fetch(eid, "(RFC822)")
                if typ != "OK" or not msg_data or not msg_data[0]:
                    continue
                msg = email.message_from_bytes(msg_data[0][1])
                from_addr = extract_email_address(msg.get("From", ""))
                in_reply_to = safe_decode(msg.get("In-Reply-To", "")).strip()
                references = safe_decode(msg.get("References", "")).strip()
                body = strip_quoted_reply(get_email_body(msg))
                if not is_internal_email(from_addr):
                    continue
                if is_rejection_reply(body):
                    for thread_id, action_list in pending_approvals.items():
                        if not (thread_id and (thread_id in in_reply_to or thread_id in references)):
                            continue
                        for action_data in action_list:
                            update_action_row(action_data["row"], status="Closed — No Action Required", notes=f"Closed per instruction from {from_addr}")
                            notice = f"Dear Team,\n\nPer your instruction, the following action item has been closed with no further follow-up.\n\nMeeting reference: {action_data['meeting_ref']}\nAction: {action_data['action']}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                            send_email(REPORT_RECIPIENTS, f"Closed — No Action Required — {action_data['action'][:50]}", notice, html_body=build_reply_html(notice))
                    continue
                if not is_approval_reply(body):
                    continue
                for thread_id, action_list in pending_approvals.items():
                    if not (thread_id and (thread_id in in_reply_to or thread_id in references)):
                        continue
                    for action_data in action_list:
                        reminder_count = action_data["reminder_count"] + 1
                        draft = draft_external_reminder(action_data["action"], action_data["responsible_name"], action_data["responsible"], action_data["due_date"], action_data["meeting_ref"], reminder_count, on_behalf_of=get_first_name(action_data["mom_sender"]))
                        if draft and action_data["email"] not in ["UNKNOWN", ""]:
                            if not action_data["responsible_name"]:
                                domain = action_data["email"].split("@")[-1] if "@" in action_data["email"] else ""
                                all_to = get_all_domain_emails(domain, action_data["all_participants"])
                                to_emails = all_to if all_to else [action_data["email"]]
                            else:
                                to_emails = [action_data["email"]]
                            cc_list = build_cc_for_external(action_data)
                            resp_label = action_data["responsible_name"] or action_data["responsible"]
                            reminder_tag = {1: "Follow-up", 2: "Second Follow-up", 3: "Escalation Notice"}.get(reminder_count, "Follow-up")
                            html_reminder = build_external_reminder_html(draft, action_data["meeting_ref"], action_data["action"], resp_label, action_data["due_date"], reminder_tag, on_behalf_of=get_first_name(action_data["mom_sender"]))
                            sent = send_email(to_emails, f"Action Item Follow-up — {action_data['meeting_ref']}", draft, html_body=html_reminder, cc_emails=cc_list)
                            if sent:
                                update_action_row(action_data["row"], status="Reminded", last_reminded=datetime.now().strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_count, draft_sent=datetime.now().strftime("%d.%m.%Y %H:%M"))
            except (imaplib.IMAP4.abort, OSError, EOFError):
                safe_logout(mail)
                try:
                    mail = get_imap_connection(); mail.select("INBOX")
                except Exception:
                    break
                continue
            except Exception as e:
                logger.error(f"Approval check error: {e}")
                continue
        safe_logout(mail)
    except Exception as e:
        logger.error(f"Action approval check error: {e}")
        if mail:
            safe_logout(mail)


def check_external_action_reminders():
    try:
        sheet = get_action_tracker_sheet()
        if not sheet:
            return
        all_values = sheet.get_all_values()
        today = datetime.now()
        for i, row in enumerate(all_values[1:], start=2):
            try:
                if len(row) < 7:
                    continue
                data = get_action_data_from_row(row)
                if data["status"] not in ["Open", "Reminded"] or not data["email"] or data["email"] in ["UNKNOWN", ""]:
                    continue
                try:
                    logged_date = datetime.strptime(data["date"], "%d.%m.%Y %H:%M")
                except:
                    continue
                days_open = (today - logged_date).days
                resp_label = data["responsible_name"] or data["responsible"]
                if days_open >= AUTO_CLOSE_DAYS:
                    update_action_row(i, status="Closed — No Response", notes=f"Auto-closed after {days_open} days")
                    notice = f"Dear {get_first_name(data['mom_sender'])},\n\nThe following action item has been automatically closed after {AUTO_CLOSE_DAYS} days with no response from {resp_label}.\n\nMeeting reference: {data['meeting_ref']}\nAction: {data['action']}\nResponsible: {resp_label} ({data['email']})\nDays open: {days_open}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    cc = [r for r in REPORT_RECIPIENTS if r.lower() != data["mom_sender"].lower()]
                    send_email([data["mom_sender"]], f"External Action Auto-Closed — {data['action'][:50]}", notice, html_body=build_reply_html(notice), cc_emails=cc)
                    continue
                reminder_due = None
                if days_open >= REMINDER_3_DAYS and data["reminder_count"] < 3:
                    reminder_due = 3
                elif days_open >= REMINDER_2_DAYS and data["reminder_count"] < 2:
                    reminder_due = 2
                elif days_open >= REMINDER_1_DAYS and data["reminder_count"] < 1:
                    reminder_due = 1
                if not reminder_due:
                    continue
                if data["last_reminded"]:
                    try:
                        last_date = datetime.strptime(data["last_reminded"], "%d.%m.%Y %H:%M")
                        if (today - last_date).days < 3:
                            continue
                    except:
                        pass
                tone_label = {1: "Follow-up", 2: "Second Follow-up", 3: "Escalation Notice"}[reminder_due]
                draft = draft_external_reminder(data["action"], data["responsible_name"], data["responsible"], data["due_date"], data["meeting_ref"], reminder_due, on_behalf_of=get_first_name(data["mom_sender"]))
                if draft:
                    if not data["responsible_name"]:
                        domain = data["email"].split("@")[-1] if "@" in data["email"] else ""
                        all_to = get_all_domain_emails(domain, data["all_participants"])
                        to_preview = ", ".join(all_to) if all_to else data["email"]
                    else:
                        to_preview = data["email"]
                    cc_preview = build_cc_for_external(data)
                    approval = f"Dear {get_first_name(data['mom_sender'])},\n\nThe action item below from {data['meeting_ref']} has been open for {days_open} days without a response from {resp_label}.\n\nAction: {data['action']}\nResponsible: {resp_label} ({data['email']})\nDue date: {data['due_date']}\n\nI have prepared a {tone_label.lower()} for your approval, or no need to close this item.\n\nWhen approved this email will be sent:\nTo: {to_preview}\nCC: {', '.join(cc_preview)}\n\n{'='*50}\nDRAFT — {tone_label.upper()} TO {resp_label.upper()}:\n{'='*50}\n\n{draft}\n\n{'='*50}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    cc = [r for r in REPORT_RECIPIENTS if r.lower() != data["mom_sender"].lower()]
                    sent = send_email([data["mom_sender"]], f"Approval Required — {tone_label} to {resp_label}", approval, html_body=build_reply_html(approval), cc_emails=cc, reply_to_msg_id=data["thread_id"], references=data["thread_id"])
                    if sent:
                        update_action_row(i, status="Draft Pending", last_reminded=today.strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_due)
            except Exception as e:
                logger.error(f"External reminder row error: {e}")
                continue
    except Exception as e:
        logger.error(f"External reminder check error: {e}")


def check_ncr_reminders():
    try:
        sheet = get_ncr_tracker_sheet()
        if not sheet:
            return
        all_values = sheet.get_all_values()
        today = datetime.now()
        for i, row in enumerate(all_values[1:], start=2):
            try:
                if len(row) < 7:
                    continue
                data = get_ncr_data_from_row(row)
                if data["status"] not in ["Open", "Reminded"] or not data["all_emails"]:
                    continue
                try:
                    logged_date = datetime.strptime(data["date"], "%d.%m.%Y %H:%M")
                except:
                    continue
                days_open = (today - logged_date).days
                resp_label = data["responsible_name"] or data["contractor"]
                reminder_count = data["reminder_count"]
                reminder_due = None
                if days_open >= REMINDER_1_DAYS and reminder_count == 0:
                    reminder_due = 1
                elif days_open >= REMINDER_2_DAYS and reminder_count == 1:
                    reminder_due = 2
                elif days_open >= REMINDER_3_DAYS and reminder_count == 2:
                    reminder_due = 3
                elif days_open >= REMINDER_3_DAYS and reminder_count >= 3:
                    if data["last_reminded"]:
                        try:
                            last_date = datetime.strptime(data["last_reminded"], "%d.%m.%Y %H:%M")
                            if (today - last_date).days >= 7:
                                reminder_due = reminder_count + 1
                        except:
                            pass
                if not reminder_due:
                    continue
                if data["last_reminded"]:
                    try:
                        last_date = datetime.strptime(data["last_reminded"], "%d.%m.%Y %H:%M")
                        if (today - last_date).days < 3:
                            continue
                    except:
                        pass
                tone = "polite and professional" if reminder_due <= 1 else "firm and urgent" if reminder_due == 2 else "formal escalation"
                prompt = f"""Draft a {tone} reminder email to a contractor regarding an open Non-Conformance Report.

NCR reference: {data['ncr_number']}
Description: {data['description']}
Contractor: {resp_label}
Date raised: {data['date_raised']}
Days open: {days_open}

Start with Dear {resp_label if not data['responsible_name'] else data['responsible_name'].split()[0]},
State clearly this NCR remains open and will not be closed until a corrective action report is submitted and accepted.
Write complete formal professional email. No bullet points or symbols.
End with:
Alex Rivera
Construction Expert
SCOPE Consulting MMC
internal@scope-iq.io"""
                response = anthropic_client.messages.create(model=MODEL, max_tokens=700, system=SYSTEM_PROMPT, messages=[{"role": "user", "content": prompt}])
                draft = response.content[0].text
                approval = f"Dear {get_first_name(data['raised_by'])},\n\nThe following NCR remains open for {days_open} days with no corrective action report received from {resp_label}.\n\nNCR reference: {data['ncr_number']}\nDescription: {data['description']}\nContractor: {resp_label} ({', '.join(data['all_emails'])})\n\nI have prepared a follow-up for your approval.\n\n{'='*50}\nDRAFT — NCR FOLLOW-UP TO {resp_label.upper()}:\n{'='*50}\n\n{draft}\n\n{'='*50}\n\nNote: This NCR will remain open until a corrective action report is received and accepted.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                cc = list(set([r for r in REPORT_RECIPIENTS if r.lower() != data["raised_by"].lower()] + data.get("client_emails", [])))
                ncr_reminder_html = build_ncr_reminder_html(approval, data["ncr_number"], data["description"], resp_label, data["date_raised"], days_open)
                sent = send_email([data["raised_by"]], f"Approval Required — NCR Follow-up to {resp_label}", approval, html_body=ncr_reminder_html, cc_emails=cc, reply_to_msg_id=data["thread_id"], references=data["thread_id"])
                if sent:
                    update_ncr_row(i, status="Reminded", last_reminded=today.strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_due)
            except Exception as e:
                logger.error(f"NCR reminder row error: {e}")
                continue
    except Exception as e:
        logger.error(f"NCR reminder check error: {e}")


def check_thread_replies(thread_ids):
    replied_threads = set()
    if not thread_ids:
        return replied_threads
    mail = None
    try:
        mail = get_imap_connection()
        mail.select("INBOX")
        typ, data = mail.search(None, "ALL")
        if typ != "OK" or not data or not data[0]:
            safe_logout(mail)
            return replied_threads
        for eid in data[0].split()[-100:]:
            try:
                typ, msg_data = mail.fetch(eid, "(RFC822)")
                if typ != "OK" or not msg_data or not msg_data[0]:
                    continue
                msg = email.message_from_bytes(msg_data[0][1])
                in_reply_to = safe_decode(msg.get("In-Reply-To", "")).strip()
                references = safe_decode(msg.get("References", "")).strip()
                sender = extract_email_address(msg.get("From", ""))
                if ZOHO_EMAIL.lower() in sender:
                    continue
                for tid in thread_ids:
                    if tid and (tid in in_reply_to or tid in references):
                        replied_threads.add(tid)
            except (imaplib.IMAP4.abort, OSError, EOFError):
                safe_logout(mail)
                try:
                    mail = get_imap_connection(); mail.select("INBOX")
                except Exception:
                    break
                continue
            except:
                continue
        safe_logout(mail)
    except Exception as e:
        logger.error(f"Check replies error: {e}")
        if mail:
            safe_logout(mail)
    return replied_threads


def check_followup_reminders():
    try:
        sheet = get_sheet("Sheet1")
        if not sheet:
            return
        all_values = sheet.get_all_values()
        if not all_values or len(all_values) < 2:
            return
        today = datetime.now()
        monitoring_threads = {}
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) >= 6 and row[5].strip() == "Monitoring":
                tid = row[6].strip() if len(row) > 6 else ""
                if tid:
                    monitoring_threads[tid] = i
        replied = check_thread_replies(set(monitoring_threads.keys()))
        for tid, row_num in monitoring_threads.items():
            if tid in replied:
                update_row(row_num, status="Closed")
        all_values = sheet.get_all_values()
        for i, row in enumerate(all_values[1:], start=2):
            try:
                if len(row) < 6:
                    continue
                date_str = row[0].strip() if row[0] else ""
                sender = row[1].strip() if len(row) > 1 else ""
                subject = row[2].strip() if len(row) > 2 else ""
                summary = row[3].strip() if len(row) > 3 else ""
                action = row[4].strip() if len(row) > 4 else ""
                status = row[5].strip() if len(row) > 5 else ""
                last_reminded = row[7].strip() if len(row) > 7 else ""
                reminder_count = int(row[8].strip()) if len(row) > 8 and row[8].strip().isdigit() else 0

                if status not in ["Open", "Monitoring"]:
                    continue
                if not sender or "@" not in sender:
                    continue
                if not any(t in sender for t in SCOPE_TEAM_EMAILS):
                    continue
                try:
                    logged_date = datetime.strptime(date_str, "%d.%m.%Y %H:%M")
                except:
                    continue
                days_open = (today - logged_date).days
                if days_open >= AUTO_CLOSE_DAYS:
                    update_row(i, status="Closed — No Response")
                    notice = f"Dear Team,\n\nThe following email thread has been automatically closed after {AUTO_CLOSE_DAYS} days.\n\nSubject: {subject}\nFrom: {sender}\nDays open: {days_open}\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    send_email(REPORT_RECIPIENTS, f"Auto-Closed — {subject}", notice, html_body=build_reply_html(notice))
                    continue
                reminder_due = None
                if days_open >= REMINDER_3_DAYS and reminder_count < 3:
                    reminder_due = 3
                elif days_open >= REMINDER_2_DAYS and reminder_count < 2:
                    reminder_due = 2
                elif days_open >= REMINDER_1_DAYS and reminder_count < 1:
                    reminder_due = 1
                if not reminder_due:
                    continue
                if last_reminded:
                    try:
                        last_date = datetime.strptime(last_reminded, "%d.%m.%Y %H:%M")
                        if (today - last_date).days < 3:
                            continue
                    except:
                        pass
                first_name = get_first_name(sender)
                if reminder_due == 1:
                    subject_line = f"Follow-up — {subject}"
                    body = f"Dear {first_name},\n\nI am writing to follow up on the email below, open for {days_open} days.\n\nSubject: {subject}\nDate raised: {date_str}\n\nSummary:\n{summary}\n\nAction required:\n{action}\n\nPlease review and respond at your earliest convenience.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    cc = [ALWAYS_CC] if sender.lower() != ALWAYS_CC.lower() else []
                elif reminder_due == 2:
                    subject_line = f"Second Follow-up — {subject}"
                    body = f"Dear {first_name},\n\nSecond follow-up. This matter has been open for {days_open} days without a response.\n\nSubject: {subject}\nDate raised: {date_str}\n\nSummary:\n{summary}\n\nAction required:\n{action}\n\nThis requires your urgent attention.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    cc = [ALWAYS_CC] if sender.lower() != ALWAYS_CC.lower() else []
                elif reminder_due == 3:
                    subject_line = f"Escalation Notice — {subject}"
                    body = f"Dear {first_name},\n\nFormal escalation. This matter has been open for {days_open} days despite two previous reminders.\n\nSubject: {subject}\nDate raised: {date_str}\n\nSummary:\n{summary}\n\nAction required:\n{action}\n\nIf no response within 7 days this will be auto-closed. Copied to management.\n\nKind regards,\n\nAlex Rivera\nConstruction Expert\nSCOPE Consulting MMC\ninternal@scope-iq.io"
                    cc = [r for r in REPORT_RECIPIENTS if r.lower() != sender.lower()]
                sent = send_email([sender], subject_line, body, cc_emails=cc, html_body=build_reply_html(body))
                if sent:
                    update_row(i, last_reminded=today.strftime("%d.%m.%Y %H:%M"), reminder_count=reminder_due)
            except Exception as e:
                logger.error(f"Reminder row error: {e}")
                continue
    except Exception as e:
        logger.error(f"Follow-up check error: {e}")


def safe_decode(value, fallback=""):
    if value is None:
        return fallback
    try:
        parts = decode_header(str(value))
        result = ""
        for part, enc in parts:
            if isinstance(part, bytes):
                result += part.decode(enc or "utf-8", errors="replace")
            else:
                result += str(part)
        return result
    except:
        return str(value) if value else fallback


def extract_email_address(header_value):
    if not header_value:
        return ""
    header_str = safe_decode(header_value)
    if "<" in header_str and ">" in header_str:
        start = header_str.rfind("<") + 1
        end = header_str.rfind(">")
        return header_str[start:end].strip().lower()
    return header_str.strip().lower()


def extract_all_emails(header_value):
    return [r["email"] for r in extract_all_emails_with_names(header_value)]


def extract_attachments(msg):
    attachments = []
    try:
        for part in msg.walk():
            content_disposition = str(part.get("Content-Disposition", ""))
            filename = part.get_filename()
            if not filename and "attachment" not in content_disposition:
                continue
            filename = safe_decode(filename, "attachment")
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            ext = filename.lower().split(".")[-1] if "." in filename else ""
            if ext in ["xlsx", "xls"]:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(payload))
                    text = f"Excel: {filename}\nSheets: {', '.join(wb.sheetnames)}\n\n"
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        text += f"\n{'='*40}\nSHEET: {sheet_name}\n{'='*40}\n"
                        for row in ws.iter_rows(max_row=500, values_only=True):
                            row_data = [str(c) for c in row if c is not None]
                            if row_data:
                                text += " | ".join(row_data) + "\n"
                    attachments.append({"name": filename, "content": text[:15000]})
                except Exception as e:
                    logger.error(f"Excel error: {e}")
            elif ext == "pdf":
                try:
                    import fitz
                    doc = fitz.open(stream=payload, filetype="pdf")
                    text = f"PDF: {filename}\n" + "".join(p.get_text() for p in doc)
                    doc.close()
                    attachments.append({"name": filename, "content": text[:5000]})
                except Exception as e:
                    logger.error(f"PDF error: {e}")
            elif ext in ["docx", "doc"]:
                try:
                    import docx
                    document = docx.Document(io.BytesIO(payload))
                    text = f"Word: {filename}\n" + "\n".join(p.text for p in document.paragraphs)
                    attachments.append({"name": filename, "content": text[:5000]})
                except Exception as e:
                    logger.error(f"Word error: {e}")
    except Exception as e:
        logger.error(f"Attachment error: {e}")
    return attachments


def get_email_body(msg):
    body = ""
    try:
        if msg.is_multipart():
            for part in msg.walk():
                try:
                    if part.get_content_type() == "text/plain" and not part.get_filename():
                        payload = part.get_payload(decode=True)
                        if payload:
                            body += payload.decode(part.get_content_charset() or "utf-8", errors="replace")
                except:
                    continue
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(msg.get_content_charset() or "utf-8", errors="replace")
    except Exception as e:
        logger.error(f"Body error: {e}")
    return body[:6000]


def send_email(to_emails, subject, body, reply_to_msg_id=None, references=None, cc_emails=None, html_body=None):
    try:
        resend.api_key = RESEND_API_KEY
        if isinstance(to_emails, str):
            to_emails = [to_emails]
        to_emails = [e for e in to_emails if e.lower() != ZOHO_EMAIL.lower()]
        if not to_emails:
            return False
        params = {"from": f"Alex Rivera <{ZOHO_EMAIL}>", "to": to_emails, "subject": subject, "text": body, "headers": {}}
        if html_body:
            params["html"] = html_body
        if cc_emails:
            cc_emails = [e for e in cc_emails if e.lower() != ZOHO_EMAIL.lower()]
            if cc_emails:
                params["cc"] = cc_emails
        if reply_to_msg_id:
            params["headers"]["In-Reply-To"] = reply_to_msg_id
            params["headers"]["References"] = references or reply_to_msg_id
        resend.Emails.send(params)
        return True
    except Exception as e:
        logger.error(f"Resend error: {e}")
        return False


def analyse_email(sender, subject, body, attachments=None, memory_files=None, is_cc=False):
    try:
        active_files = attachments if attachments else memory_files
        using_memory = not attachments and bool(memory_files)
        if is_cc:
            full_content = body
            if active_files:
                full_content += "\n\nFILES:\n" + "\n".join(f"{f['name']}:\n{f['content']}" for f in active_files)
            prompt = f"""CC'd email — internal analysis only. Do not reply.
From: {sender}
Subject: {subject}
Content: {full_content}

Full internal assessment:
1. Email type.
2. Summary in three to five sentences.
3. Commercial, technical, contractual or programme implications.
4. Specific actions — what, who, by when.
5. Risk level: High, Medium or Low.
6. Recommended response deadline.
Plain professional prose. Numbered paragraphs. No symbols."""
        else:
            if active_files:
                memory_note = ""
                if using_memory:
                    saved_at = active_files[0].get("saved_at", "previously")
                    file_names = ", ".join(f["name"] for f in active_files)
                    memory_note = f"Note: Using files previously received ({file_names}, saved {saved_at}).\n\n"
                files_content = "\n".join(f"{f['name']}:\n{f['content']}" for f in active_files)
                prompt = f"""Email from SCOPE team member.
From: {sender}
Subject: {subject}
Email body: {body}

{memory_note}FILE CONTENT:
{files_content}

Analyse every discipline, every sheet, every line item equally.
For each item: description, quantity, unit rate, market assessment, Baku market range, quantity concerns.
Flag missing scope. Discipline risk summary. Total risk exposure at end.
Write complete formal professional reply now."""
            else:
                prompt = f"""Email from SCOPE team member.
From: {sender}
Subject: {subject}
Content: {body}
Write complete formal professional reply. If files needed and never provided, request them."""
        response = anthropic_client.messages.create(model=MODEL, max_tokens=4000, system=SYSTEM_PROMPT, messages=[{"role": "user", "content": prompt}])
        return response.content[0].text
    except Exception as e:
        logger.error(f"Analysis error: {e}")
        return None


def build_reply_html(body_text):
    today = datetime.now().strftime("%d %B %Y")
    time_now = datetime.now().strftime("%H:%M")
    paragraphs = body_text.strip().split("\n\n")
    html_body = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if "Alex Rivera" in para and "SCOPE Consulting" in para:
            lines = para.split("\n")
            sig_html = ""
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "Alex Rivera" in line:
                    sig_html += f'<div style="font-size:13px;font-weight:600;color:#1a2942;">{line}</div>'
                elif "internal@scope-iq.io" in line:
                    sig_html += f'<div style="font-size:12px;color:#3CB496;">{line}</div>'
                else:
                    sig_html += f'<div style="font-size:12px;color:#666;">{line}</div>'
            html_body += f'<div style="margin-top:24px;padding-top:16px;border-top:1px solid #f0f0f0;line-height:1.8;">{sig_html}</div>'
        else:
            lines = para.split("\n")
            para_html = "<br>".join(line.strip() for line in lines if line.strip())
            html_body += f'<p style="font-size:14px;color:#333;line-height:1.8;margin:0 0 16px;">{para_html}</p>'
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:620px;margin:0 auto;padding:20px 0;">
  <div style="background:#1a2942;border-radius:12px 12px 0 0;padding:18px 28px;">
    <div style="display:flex;justify-content:space-between;align-items:center;">
      <div style="color:#fff;font-size:20px;font-weight:600;letter-spacing:1px;">SCOPE <span style="color:#3CB496;">IQ</span></div>
      <div style="font-size:11px;color:#8facc8;">{today} &nbsp;·&nbsp; {time_now} Baku</div>
    </div>
    <div style="font-size:12px;color:#8facc8;margin-top:6px;">Response from Alex Rivera &nbsp;·&nbsp; Construction Expert</div>
  </div>
  <div style="background:#fff;border:1px solid #e8e8e8;border-top:none;border-radius:0 0 12px 12px;padding:28px 28px 24px;">
    {html_body}
    <div style="background:#f8f9fa;border-radius:6px;padding:10px 14px;margin-top:20px;"><div style="font-size:11px;color:#888;">This response was prepared by <strong style="color:#1a2942;">Alex Rivera</strong>, using <strong style="color:#3CB496;">SCOPE IQ</strong>.</div></div>
  </div>
</div></body></html>"""


def process_emails():
    load_processed_ids()
    mail = None
    try:
        mail = get_imap_connection()
        mail.select("INBOX")
        typ, data = mail.search(None, "ALL")
        if typ != "OK" or not data or not data[0]:
            safe_logout(mail)
            return
        all_ids = data[0].split()
        if not all_ids:
            safe_logout(mail)
            return
        recent = all_ids[-50:]

        for eid in reversed(recent):
            try:
                eid_str = eid.decode() if isinstance(eid, bytes) else str(eid)
                try:
                    typ, msg_data = mail.fetch(eid, "(RFC822)")
                    if typ != "OK" or not msg_data or not msg_data[0]:
                        continue
                except (imaplib.IMAP4.abort, OSError, EOFError):
                    safe_logout(mail)
                    try:
                        mail = get_imap_connection(); mail.select("INBOX")
                        typ, msg_data = mail.fetch(eid, "(RFC822)")
                        if typ != "OK" or not msg_data or not msg_data[0]:
                            continue
                    except Exception:
                        break

                raw = msg_data[0][1]
                if not raw:
                    continue
                msg = email.message_from_bytes(raw)
                msg_id_hdr = safe_decode(msg.get("Message-ID"), "").strip()
                unique_id = msg_id_hdr or eid_str
                if is_processed(unique_id):
                    continue
                mark_as_processed(unique_id)

                sender = extract_email_address(msg.get("From", ""))
                subject = safe_decode(msg.get("Subject"), "No subject")
                to_field = safe_decode(msg.get("To"), "").lower()
                cc_field = safe_decode(msg.get("CC"), "").lower()
                references = safe_decode(msg.get("References"), "")
                in_reply_to = safe_decode(msg.get("In-Reply-To", "")).strip()

                to_with_names = extract_all_emails_with_names(msg.get("To", ""))
                cc_with_names = extract_all_emails_with_names(msg.get("CC", ""))
                from_with_name = extract_email_with_name(msg.get("From", ""))
                to_addresses = [p["email"] for p in to_with_names]
                cc_addresses = [p["email"] for p in cc_with_names]

                if ZOHO_EMAIL.lower() in sender:
                    continue

                is_direct = ZOHO_EMAIL.lower() in to_field
                is_cc_email = ZOHO_EMAIL.lower() in cc_field
                is_internal = any(t in sender for t in SCOPE_TEAM_EMAILS)
                is_external = not is_internal and not is_internal_email(sender)

                if not is_direct and not is_cc_email:
                    continue

                body_raw = get_email_body(msg)
                body_clean = strip_quoted_reply(body_raw)
                attachments = extract_attachments(msg)

                if attachments:
                    save_files_to_memory(sender, attachments)
                    memory_files = None
                else:
                    memory_files = load_files_from_memory(sender)

                if is_cc_email:
                    if is_internal and is_ncr_email(subject, body_raw, attachments):
                        all_with_names = []
                        seen = set()
                        for p in [from_with_name] + to_with_names + cc_with_names:
                            if p["email"] and p["email"] not in seen:
                                seen.add(p["email"]); all_with_names.append(p)
                        for se in extract_emails_from_text(body_raw):
                            if se not in seen and se != ZOHO_EMAIL.lower():
                                seen.add(se); all_with_names.append({"email": se, "name": ""})
                        process_ncr_email(sender, subject, body_raw, attachments, all_with_names, msg_id_hdr)
                    elif is_internal and is_mom_email(subject, body_raw, attachments):
                        all_with_names = []
                        seen = set()
                        for p in [from_with_name] + to_with_names + cc_with_names:
                            if p["email"] and p["email"] not in seen:
                                seen.add(p["email"]); all_with_names.append(p)
                        for se in extract_emails_from_text(body_raw):
                            if se not in seen and se != ZOHO_EMAIL.lower():
                                seen.add(se); all_with_names.append({"email": se, "name": ""})
                        process_mom_email(sender, subject, body_raw, attachments, all_with_names, msg_id_hdr)
                    else:
                        analysis = analyse_email(sender, subject, body_clean, attachments=attachments, memory_files=memory_files, is_cc=True)
                        if analysis:
                            save_to_monitoring(sender, subject, analysis[:400], analysis[:500], msg_id_hdr, "Monitoring")

                elif is_direct and is_internal:
                    if is_daily_report_reply(subject):
                        if sender in SCOPE_TEAM_EMAILS:
                            handle_daily_report_reply(sender, body_clean, msg_id_hdr, references)
                        else:
                            logger.info(f"Daily report reply from {sender} — not in SCOPE_TEAM_EMAILS, ignoring")
                            save_to_monitoring(sender, subject, f"Daily report reply received from unauthorised sender {sender}", "No action taken — sender not in authorised team list", msg_id_hdr, "Monitoring")
                        continue

                    thread_actions = find_all_open_actions_matching_refs(in_reply_to, references)
                    if not thread_actions:
                        thread_actions = find_open_actions_by_subject(subject)

                    ncr_matches = find_all_open_ncrs_matching_refs(in_reply_to, references)
                    if not ncr_matches:
                        ncr_matches = find_open_ncrs_by_subject(subject)

                    if thread_actions:
                        draft_pending = [a for a in thread_actions if a["status"] == "Draft Pending"]
                        if draft_pending and is_approval_reply(body_clean) and not is_rejection_reply(body_clean):
                            for action_data in draft_pending:
                                dispatch_approved_external_draft(action_data)
                        else:
                            handle_mom_thread_reply(sender, body_clean, thread_actions, msg_id_hdr, in_reply_to, references)
                    elif ncr_matches:
                        handle_ncr_thread_reply(sender, body_clean, ncr_matches, msg_id_hdr, references)
                    elif is_short_reply(body_clean):
                        all_open_ncrs = find_all_recent_open_ncrs()
                        if all_open_ncrs:
                            handle_ncr_thread_reply(sender, body_clean, all_open_ncrs, msg_id_hdr, references)
                        else:
                            save_to_monitoring(sender, subject, f"Short reply received but no matching open NCR/MOM/action found: '{body_clean[:100]}'", "Please clarify which item this reply relates to", msg_id_hdr, "Monitoring")
                    elif is_ncr_email(subject, body_raw, attachments):
                        all_with_names = []
                        seen = set()
                        for p in [from_with_name] + to_with_names + cc_with_names:
                            if p["email"] and p["email"] not in seen:
                                seen.add(p["email"]); all_with_names.append(p)
                        for se in extract_emails_from_text(body_raw):
                            if se not in seen and se != ZOHO_EMAIL.lower():
                                seen.add(se); all_with_names.append({"email": se, "name": ""})
                        process_ncr_email(sender, subject, body_raw, attachments, all_with_names, msg_id_hdr)
                        save_to_memory(sender, subject, "NCR processed — see NCR Tracker", "Review NCR Tracker", "Closed")
                    elif is_mom_email(subject, body_raw, attachments):
                        all_with_names = []
                        seen = set()
                        for p in [from_with_name] + to_with_names + cc_with_names:
                            if p["email"] and p["email"] not in seen:
                                seen.add(p["email"]); all_with_names.append(p)
                        for se in extract_emails_from_text(body_raw):
                            if se not in seen and se != ZOHO_EMAIL.lower():
                                seen.add(se); all_with_names.append({"email": se, "name": ""})
                        process_mom_email(sender, subject, body_raw, attachments, all_with_names, msg_id_hdr)
                        save_to_memory(sender, subject, "MOM processed — see Action Tracker", "Review Action Tracker", "Closed")
                    else:
                        analysis = analyse_email(sender, subject, body_clean, attachments=attachments, memory_files=memory_files, is_cc=False)
                        if analysis:
                            reply_sub = f"Re: {subject}" if not subject.startswith("Re:") else subject
                            all_recipients = list(set([sender] + [a for a in to_addresses if a != ZOHO_EMAIL.lower()] + [a for a in cc_addresses if a != ZOHO_EMAIL.lower()]))
                            new_references = f"{references} {msg_id_hdr}".strip() if references else msg_id_hdr
                            sent = send_email(all_recipients, reply_sub, analysis, reply_to_msg_id=msg_id_hdr, references=new_references, html_body=build_reply_html(analysis))
                            save_to_memory(sender, subject, analysis[:400], "Replied by Alex", "Closed" if sent else "Open")

                elif is_direct and is_external:
                    matched_action = find_action_by_thread(in_reply_to, references)
                    if matched_action:
                        route_external_reply_for_approval(sender, subject, body_clean, matched_action, msg_id_hdr)
                    else:
                        save_to_monitoring(sender, subject, f"External email received from {sender}. No matching tracked action found.", "Review if action required", msg_id_hdr, "Monitoring")

            except Exception as e:
                logger.error(f"Email error: {e}")
                continue

        safe_logout(mail)
    except imaplib.IMAP4.error as e:
        logger.error(f"IMAP auth error: {e}")
        if mail:
            safe_logout(mail)
    except Exception as e:
        logger.error(f"IMAP error: {e}")
        if mail:
            safe_logout(mail)


def send_morning_report():
    try:
        pending, closed = read_memory_for_report()
        open_actions, _, flagged = read_actions_for_report()
        open_ncrs, _ = read_ncrs_for_report()
        today = datetime.now().strftime("%d %B %Y")
        day_name = datetime.now().strftime("%A")
        time_now = datetime.now().strftime("%H:%M")
        html_body = build_report_html(pending, closed, today, day_name, time_now, open_actions=open_actions, flagged=flagged, open_ncrs=open_ncrs)
        total_open = len(pending) + len(open_actions) + len(open_ncrs)
        plain = f"SCOPE IQ Daily Report — {today}\nInternal: {len(pending)} | External: {len(open_actions)} | Open NCRs: {len(open_ncrs)}\nAlex Rivera | SCOPE Consulting MMC"
        subject_line = f"SCOPE IQ Daily Report — {today}" + (f" — {total_open} Open Item(s)" if total_open else " — All Clear")
        send_email(REPORT_RECIPIENTS, subject_line, plain, html_body=html_body)
    except Exception as e:
        logger.error(f"Morning report error: {e}")


def main():
    logger.info("Alex Email Service starting")
    load_processed_ids()

    schedule.every(10).minutes.do(process_emails)
    schedule.every().day.at("05:00").do(send_morning_report)
    schedule.every(6).hours.do(check_followup_reminders)
    schedule.every(6).hours.do(check_external_action_replies)
    schedule.every(6).hours.do(check_action_approvals)
    schedule.every(6).hours.do(check_external_action_reminders)
    schedule.every(6).hours.do(check_mom_confirmation_and_rejections)
    schedule.every(6).hours.do(check_ncr_reminders)

    process_emails()

    while True:
        schedule.run_pending()
        time.sleep(30)


if __name__ == "__main__":
    main()
